import argparse
import io
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from contextlib import redirect_stdout
from unittest.mock import Mock, patch

import requests
from openpyxl import load_workbook

import scripts.collect_it_jobs as job_collector

from scripts.collect_it_jobs import (
    CHAT_ID,
    CollectionResult,
    FALLBACK_SOURCES,
    JOB_HEADERS,
    build_summary_message,
    build_workbook,
    classify_jobkorea_detail,
    classify_rallit_position,
    classify_saramin_detail,
    classify_wanted_job,
    collect_jumpit_jobs,
    collect_rallit_jobs,
    collect_starting_jobs,
    collect_wanted_jobs,
    delivery_already_sent,
    delivery_key,
    explicit_year_floor_at_least,
    filter_previous_day_duplicates,
    load_research_ledger,
    listing_has_backend_signal,
    merge_platform_links,
    main,
    normalize_research_url,
    parse_args,
    parse_source_names,
    platform_search_terms,
    record_delivery,
    record_researched_url,
    save_research_ledger,
    select_jobs_for_workbook,
    send_telegram,
    scheduled_sources,
    validate_workbook_for_delivery,
    worknet_detail_fields,
)


class CollectItJobsTest(unittest.TestCase):
    def test_listing_backend_signal_rejects_generic_big_data_ai_category(self):
        item = {
            "RecruitTitle": "총무지원팀 계약직 채용",
            "Depth": "빅데이터/AI, 사무/총무/법무",
            "AssignedTaskNameListString": "총무지원팀",
        }

        self.assertFalse(listing_has_backend_signal(item, ["RecruitTitle", "Depth", "AssignedTaskNameListString"]))
        self.assertFalse(listing_has_backend_signal({"title": "인프라팀 전기 담당자 채용"}, ["title"]))

    def test_worknet_detail_fields_reads_summary_list_and_deadline(self):
        html = """
        <p class="corp_info"><strong>테스트회사</strong></p>
        <strong class="title">백엔드 개발자</strong>
        <li><em class="tit">경력</em><p>신입</p></li>
        <li><em class="tit">지역</em><p>부산</p></li>
        <li><em class="tit">고용형태</em><p>정규직</p></li>
        <strong>접수 마감일</strong><p class="b1_sb">2026.08.25 24:00</p>
        <div class="fold">직무내용 Spring API 개발</div>
        """

        self.assertEqual(
            worknet_detail_fields(html),
            ("테스트회사", "백엔드 개발자", "신입", "부산", "2026.08.25 24:00", "정규직", "직무내용 Spring API 개발"),
        )

    def test_collect_starting_jobs_does_not_use_company_intro_as_primary_job_signal(self):
        item = {
            "publicNumber": 1,
            "companyName": "테스트회사",
            "title": "ML Engineer 채용",
            "companyIntro": "백엔드 개발자가 있는 회사입니다.",
        }

        with patch("scripts.collect_it_jobs.fetch_json", return_value={"rows": [item]}):
            result = collect_starting_jobs(object(), "2026-07-30", "2026-07-30 09:00 KST", 1, set())

        self.assertEqual(result.checked, 0)
        self.assertEqual(result.included, [])

    def test_experience_floor_does_not_treat_company_age_as_required_career(self):
        self.assertFalse(explicit_year_floor_at_least("업력 6년차 스타트업", 5))
        self.assertTrue(explicit_year_floor_at_least("경력 6년차 개발자", 5))

    def test_fallback_sources_add_five_detail_verifiable_platforms(self):
        self.assertEqual(
            FALLBACK_SOURCES,
            ("jumpit", "catch", "incruit", "worknet", "starting"),
        )

    def test_collect_jumpit_jobs_checks_public_detail_before_including(self):
        list_item = {
            "id": 101,
            "title": "백엔드 개발자",
            "companyName": "테스트회사",
            "techStacks": ["Java", "Spring Boot"],
            "minCareer": 1,
            "maxCareer": 3,
            "locations": ["대전"],
            "closedAt": "2026-08-30T23:59:59",
        }
        detail = {
            **list_item,
            "responsibility": "Spring 기반 API 개발 및 운영",
            "qualifications": "Java와 Spring을 사용한 개발 경력 1년 이상",
            "preferredRequirements": "AWS 경험",
            "employmentType": "정규직",
        }

        def fake_fetch_json(_session, url, params=None):
            if "positions" in url:
                return {"result": {"positions": [list_item]}}
            return {"result": detail}

        with patch("scripts.collect_it_jobs.fetch_json", side_effect=fake_fetch_json):
            result = collect_jumpit_jobs(object(), "2026-07-30", "2026-07-30 09:00 KST", 1, set())

        self.assertEqual(result.checked, 1)
        self.assertEqual(len(result.included), 1)
        self.assertEqual(result.included[0]["회사명"], "테스트회사")
        self.assertEqual(result.included[0]["주요업무 요약"], "Spring 기반 API 개발 및 운영")
        self.assertEqual(
            parse_source_names("all"),
            ["wanted", "rallit", "saramin", "jobkorea", *FALLBACK_SOURCES],
        )

    def test_scheduled_sources_uses_three_to_one_platform_cycle(self):
        self.assertEqual(scheduled_sources("2026-07-30"), ["saramin", "jobkorea"])
        self.assertEqual(scheduled_sources("2026-07-31"), ["saramin", "jobkorea"])
        self.assertEqual(scheduled_sources("2026-08-01"), ["saramin", "jobkorea"])
        self.assertEqual(scheduled_sources("2026-08-02"), ["wanted", "rallit"])
        self.assertEqual(scheduled_sources("2026-08-03"), ["saramin", "jobkorea"])

    def test_research_ledger_round_trip_and_tracking_param_normalization(self):
        url = "https://example.com/job?rec_idx=7&utm_source=test&stext=backend"
        expected = "https://example.com/job?rec_idx=7"

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "researched.json"
            records = load_research_ledger(path)
            record_researched_url(records, url, "Saramin", "included", "2026-07-28 09:00 KST")
            save_research_ledger(path, records)
            restored = load_research_ledger(path)

        self.assertEqual(normalize_research_url(url), expected)
        self.assertEqual(restored[expected]["source"], "Saramin")
        self.assertEqual(restored[expected]["outcome"], "included")

    def test_research_ledger_normalizes_preexisting_tracking_urls_on_load(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "researched.json"
            path.write_text(
                '{"https://example.com/job?rec_idx=7&utm_source=test": {"source": "Saramin"}}',
                encoding="utf-8",
            )
            restored = load_research_ledger(path)

        self.assertEqual(set(restored), {"https://example.com/job?rec_idx=7"})

    def test_researched_wanted_candidate_is_skipped_before_detail_fetch(self):
        search_job = {"id": 1, "position": "백엔드 개발자", "annual_from": 1, "annual_to": 3, "status": "active"}

        def fake_fetch_json(_session, url, params=None):
            if url == "https://www.wanted.co.kr/api/v4/jobs":
                return {"data": [search_job]}
            self.fail(f"known candidate detail must not be fetched: {url}")

        with patch("scripts.collect_it_jobs.fetch_json", side_effect=fake_fetch_json):
            result = collect_wanted_jobs(
                object(),
                "2026-07-30",
                "2026-07-30 09:00 KST",
                200,
                {"https://www.wanted.co.kr/wd/1"},
            )

        self.assertEqual(result.checked, 0)
        self.assertEqual(result.skipped_researched, 1)

    def test_rallit_paginates_until_two_hundred_candidates(self):
        def item(position_id):
            return {
                "id": position_id,
                "title": "백엔드 개발자",
                "companyName": f"회사{position_id}",
                "jobLevels": ["BEGINNER"],
                "jobSkillKeywords": ["Java", "Spring"],
                "status": {"code": "HIRING"},
            }

        def detail(position_id):
            return {
                "id": position_id,
                "title": "백엔드 개발자",
                "companyName": f"회사{position_id}",
                "jobLevels": ["BEGINNER"],
                "jobs": [{"name": "백엔드 개발"}],
                "responsibilities": "<p>Java Spring API 개발</p>",
                "basicQualifications": "<p>서버 개발 경험</p>",
                "status": {"code": "HIRING"},
            }

        pages = {page: [item((page - 1) * 60 + index) for index in range(1, 61)] for page in range(1, 5)}

        def fake_fetch_json(_session, url, params=None):
            if url == "https://www.rallit.com/client/api/v1/position":
                return {"data": {"items": pages.get(params["pageNumber"], [])}}
            position_id = int(url.rsplit("/", 1)[-1])
            return {"data": detail(position_id)}

        with patch("scripts.collect_it_jobs.fetch_json", side_effect=fake_fetch_json):
            result = collect_rallit_jobs(object(), "2026-07-30", "2026-07-30 09:00 KST", 200, set())

        self.assertEqual(result.checked, 200)
        self.assertEqual(len(result.included), 200)

    def test_main_uses_wanted_rallit_schedule_day_and_saves_successful_research_records(self):
        included = CollectionResult(
            [{"회사명": "포함회사", "공고명": "백엔드 개발자", "출처": "Wanted", "링크": "https://example.com/included"}],
            [],
            1,
            0,
            {"https://example.com/included": {"source": "Wanted", "outcome": "included", "checked_at": "2026-07-28 09:00 KST"}},
        )
        excluded = CollectionResult(
            [],
            [],
            1,
            0,
            {"https://example.com/excluded": {"source": "Rallit", "outcome": "excluded", "checked_at": "2026-07-28 09:00 KST"}},
        )

        with tempfile.TemporaryDirectory() as tmp:
            ledger_path = Path(tmp) / "researched.json"
            args = argparse.Namespace(
                output=Path(tmp) / "it-backend-jobs-2026-07-28.xlsx",
                max_details=200,
                sources=None,
                send_telegram=False,
                chat_id=CHAT_ID,
                token_env="TELEGRAM_BOT_TOKEN",
                memory_path=None,
                research_ledger=ledger_path,
                delivery_state=Path(tmp) / "delivery.json",
                force_send=False,
                quiet=True,
            )
            with patch("scripts.collect_it_jobs.parse_args", return_value=args), patch(
                "scripts.collect_it_jobs.kst_now", return_value=datetime(2026, 8, 2, 9, 0)
            ), patch("scripts.collect_it_jobs.scan_official_sources", return_value=[]), patch(
                "scripts.collect_it_jobs.collect_wanted_jobs", return_value=included
            ) as wanted, patch("scripts.collect_it_jobs.collect_rallit_jobs", return_value=excluded) as rallit, patch(
                "scripts.collect_it_jobs.collect_saramin_jobs"
            ) as saramin, patch("scripts.collect_it_jobs.collect_jobkorea_jobs") as jobkorea:
                self.assertEqual(main(), 0)

            self.assertEqual(wanted.call_count, 1)
            self.assertEqual(rallit.call_count, 1)
            self.assertEqual(saramin.call_count, 0)
            self.assertEqual(jobkorea.call_count, 0)
            self.assertEqual(
                set(load_research_ledger(ledger_path)),
                {"https://example.com/included", "https://example.com/excluded"},
            )

    def test_main_uses_first_fallback_source_when_primary_pair_has_no_new_jobs(self):
        empty = CollectionResult([], [], 0, 0, {})
        fallback_jobs = [
            {
                "회사명": f"점핏회사{index}",
                "공고명": f"백엔드 개발자 {index}",
                "출처": "Jumpit",
                "링크": f"https://example.com/jumpit/{index}",
                "2년차 적합도": "상",
                "우선순위": "상",
            }
            for index in range(20)
        ]
        fallback = CollectionResult(fallback_jobs, [], 20, 0, {})

        with tempfile.TemporaryDirectory() as tmp:
            args = argparse.Namespace(
                output=Path(tmp) / "it-backend-jobs-2026-07-30.xlsx",
                max_details=200,
                sources=None,
                send_telegram=False,
                chat_id=CHAT_ID,
                token_env="TELEGRAM_BOT_TOKEN",
                memory_path=None,
                research_ledger=Path(tmp) / "researched.json",
                delivery_state=Path(tmp) / "delivery.json",
                force_send=False,
                quiet=True,
            )
            with patch("scripts.collect_it_jobs.parse_args", return_value=args), patch(
                "scripts.collect_it_jobs.kst_now", return_value=datetime(2026, 7, 30, 9, 0)
            ), patch("scripts.collect_it_jobs.scan_official_sources", return_value=[]), patch(
                "scripts.collect_it_jobs.collect_saramin_jobs", return_value=empty
            ), patch("scripts.collect_it_jobs.collect_jobkorea_jobs", return_value=empty), patch(
                "scripts.collect_it_jobs.collect_jumpit_jobs", return_value=fallback, create=True
            ) as jumpit, patch("scripts.collect_it_jobs.collect_catch_jobs", create=True) as catch:
                self.assertEqual(main(), 0)

        self.assertEqual(jumpit.call_count, 1)
        self.assertEqual(catch.call_count, 0)

    def test_main_runs_fallback_when_two_hundred_raw_jobs_reduce_to_one_unique_job(self):
        primary = CollectionResult(
            [
                {
                    "회사명": "중복회사",
                    "공고명": "백엔드 개발자",
                    "링크": f"https://example.com/primary/{index}",
                    "출처": "Saramin",
                    "2년차 적합도": "상",
                    "우선순위": "상",
                }
                for index in range(200)
            ],
            [],
            200,
            0,
            {},
        )
        jumpit = CollectionResult(
            [
                {
                    "회사명": f"점핏회사{index}",
                    "공고명": f"백엔드 개발자 {index}",
                    "링크": f"https://example.com/jumpit/{index}",
                    "출처": "Jumpit",
                    "2년차 적합도": "상",
                    "우선순위": "상",
                }
                for index in range(199)
            ],
            [],
            199,
            0,
            {},
        )
        empty = CollectionResult([], [], 0, 0, {})

        with tempfile.TemporaryDirectory() as tmp:
            args = argparse.Namespace(
                output=Path(tmp) / "it-backend-jobs-2026-07-30.xlsx",
                max_details=200,
                sources=None,
                send_telegram=False,
                chat_id=CHAT_ID,
                token_env="TELEGRAM_BOT_TOKEN",
                memory_path=None,
                research_ledger=Path(tmp) / "researched.json",
                delivery_state=Path(tmp) / "delivery.json",
                force_send=False,
                quiet=True,
            )
            with patch("scripts.collect_it_jobs.parse_args", return_value=args), patch(
                "scripts.collect_it_jobs.kst_now", return_value=datetime(2026, 7, 30, 9, 0)
            ), patch("scripts.collect_it_jobs.scan_official_sources", return_value=[]), patch(
                "scripts.collect_it_jobs.collect_saramin_jobs", return_value=primary
            ), patch("scripts.collect_it_jobs.collect_jobkorea_jobs", return_value=empty), patch(
                "scripts.collect_it_jobs.collect_jumpit_jobs", return_value=jumpit
            ) as collect_jumpit:
                self.assertEqual(main(), 0)
            workbook = load_workbook(args.output, data_only=True)

        self.assertEqual(collect_jumpit.call_count, 1)
        self.assertEqual(workbook["채용공고"].max_row - 1, 200)

    def test_main_stops_fallback_when_primary_has_two_hundred_unique_jobs(self):
        primary = CollectionResult(
            [
                {
                    "회사명": f"사람인회사{index}",
                    "공고명": f"백엔드 개발자 {index}",
                    "링크": f"https://example.com/saramin/{index}",
                    "출처": "Saramin",
                    "2년차 적합도": "상",
                    "우선순위": "상",
                }
                for index in range(200)
            ],
            [],
            200,
            0,
            {},
        )
        empty = CollectionResult([], [], 0, 0, {})

        with tempfile.TemporaryDirectory() as tmp:
            args = argparse.Namespace(
                output=Path(tmp) / "it-backend-jobs-2026-07-30.xlsx",
                max_details=200,
                sources=None,
                send_telegram=False,
                chat_id=CHAT_ID,
                token_env="TELEGRAM_BOT_TOKEN",
                memory_path=None,
                research_ledger=Path(tmp) / "researched.json",
                delivery_state=Path(tmp) / "delivery.json",
                force_send=False,
                quiet=True,
            )
            with patch("scripts.collect_it_jobs.parse_args", return_value=args), patch(
                "scripts.collect_it_jobs.kst_now", return_value=datetime(2026, 7, 30, 9, 0)
            ), patch("scripts.collect_it_jobs.scan_official_sources", return_value=[]), patch(
                "scripts.collect_it_jobs.collect_saramin_jobs", return_value=primary
            ), patch("scripts.collect_it_jobs.collect_jobkorea_jobs", return_value=empty), patch(
                "scripts.collect_it_jobs.collect_jumpit_jobs"
            ) as collect_jumpit:
                self.assertEqual(main(), 0)

        self.assertEqual(collect_jumpit.call_count, 0)

    def test_collect_career_jobs_opens_public_detail_before_including(self):
        self.assertTrue(hasattr(job_collector, "collect_career_jobs"))
        item = {
            "regno": 123,
            "subject": "백엔드 개발자",
            "company_name": "커리어테스트",
            "career_years": "2년",
            "area_name": "부산",
            "area_name2": "해운대구",
            "work_type_name": "정규직",
            "apply_end_dateString": "2026-08-15",
        }
        detail = """
        <html><title>커리어테스트 - 백엔드 개발자 | 커리어</title><body>
        Spring Boot 기반 API와 서버 개발 업무를 담당합니다. Java, Spring, MySQL을 사용해
        고객 서비스의 백엔드 기능을 설계하고 운영합니다. 경력 2년 이상인 개발자를 찾으며,
        REST API를 설계하고 장애를 분석해 안정적인 서비스를 제공할 수 있어야 합니다.
        동료와 코드 리뷰를 진행하고 배포 자동화와 모니터링을 함께 개선합니다.
        </body></html>
        """
        session = Mock()
        response = Mock(text=detail)
        response.raise_for_status.return_value = None
        session.get.return_value = response

        with patch(
            "scripts.collect_it_jobs.fetch_json_post", return_value={"list": [item], "Total": 1}, create=True
        ):
            result = job_collector.collect_career_jobs(
                session, "2026-07-30", "2026-07-30 09:00 KST", 1, set()
            )

        self.assertEqual(len(result.included), 1)
        self.assertEqual(result.included[0]["회사명"], "커리어테스트")
        self.assertEqual(result.included[0]["지역/근무형태"], "부산 해운대구 / 정규직")
        self.assertEqual(result.included[0]["마감일"], "2026-08-15")
        self.assertIn("/recruit/view/123", session.get.call_args.args[0])

    def test_main_runs_career_only_after_core_fallbacks_are_exhausted_below_target(self):
        empty = CollectionResult([], [], 0, 0, {})
        career = CollectionResult(
            [
                {
                    "회사명": f"커리어회사{index}",
                    "공고명": f"백엔드 개발자 {index}",
                    "링크": f"https://example.com/career/{index}",
                    "출처": "Career",
                    "2년차 적합도": "상",
                    "우선순위": "상",
                }
                for index in range(200)
            ],
            [],
            200,
            0,
            {},
        )

        with tempfile.TemporaryDirectory() as tmp:
            args = argparse.Namespace(
                output=Path(tmp) / "it-backend-jobs-2026-07-30.xlsx",
                max_details=200,
                sources=None,
                send_telegram=False,
                chat_id=CHAT_ID,
                token_env="TELEGRAM_BOT_TOKEN",
                memory_path=None,
                research_ledger=Path(tmp) / "researched.json",
                delivery_state=Path(tmp) / "delivery.json",
                force_send=False,
                quiet=True,
            )
            with patch("scripts.collect_it_jobs.parse_args", return_value=args), patch(
                "scripts.collect_it_jobs.kst_now", return_value=datetime(2026, 7, 30, 9, 0)
            ), patch("scripts.collect_it_jobs.scan_official_sources", return_value=[]), patch(
                "scripts.collect_it_jobs.collect_saramin_jobs", return_value=empty
            ), patch("scripts.collect_it_jobs.collect_jobkorea_jobs", return_value=empty), patch(
                "scripts.collect_it_jobs.collect_jumpit_jobs", return_value=empty
            ), patch("scripts.collect_it_jobs.collect_catch_jobs", return_value=empty), patch(
                "scripts.collect_it_jobs.collect_incruit_jobs", return_value=empty
            ), patch("scripts.collect_it_jobs.collect_worknet_jobs", return_value=empty), patch(
                "scripts.collect_it_jobs.collect_starting_jobs", return_value=empty
            ), patch(
                "scripts.collect_it_jobs.collect_career_jobs", return_value=career, create=True
            ) as collect_career:
                self.assertEqual(main(), 0)

        self.assertEqual(collect_career.call_count, 1)

    def test_main_continues_when_one_platform_list_request_fails(self):
        empty = CollectionResult([], [], 0, 0, {})
        fallback_jobs = [
            {
                "회사명": f"점핏회사{index}",
                "공고명": f"백엔드 개발자 {index}",
                "출처": "Jumpit",
                "링크": f"https://example.com/jumpit/{index}",
                "2년차 적합도": "상",
                "우선순위": "상",
            }
            for index in range(20)
        ]
        fallback = CollectionResult(fallback_jobs, [], 20, 0, {})

        with tempfile.TemporaryDirectory() as tmp:
            args = argparse.Namespace(
                output=Path(tmp) / "it-backend-jobs-2026-07-30.xlsx",
                max_details=200,
                sources=None,
                send_telegram=False,
                chat_id=CHAT_ID,
                token_env="TELEGRAM_BOT_TOKEN",
                memory_path=None,
                research_ledger=Path(tmp) / "researched.json",
                delivery_state=Path(tmp) / "delivery.json",
                force_send=False,
                quiet=True,
            )
            with patch("scripts.collect_it_jobs.parse_args", return_value=args), patch(
                "scripts.collect_it_jobs.kst_now", return_value=datetime(2026, 7, 30, 9, 0)
            ), patch("scripts.collect_it_jobs.scan_official_sources", return_value=[]), patch(
                "scripts.collect_it_jobs.collect_saramin_jobs", side_effect=requests.ConnectionError("reset")
            ), patch("scripts.collect_it_jobs.collect_jobkorea_jobs", return_value=empty), patch(
                "scripts.collect_it_jobs.collect_jumpit_jobs", return_value=fallback
            ):
                self.assertEqual(main(), 0)

            workbook = load_workbook(args.output, data_only=True)

        excluded_rows = list(workbook["제외_검토"].iter_rows(min_row=2, values_only=True))
        self.assertTrue(any(row[3] == "Saramin" and "목록 접근 실패" in row[4] for row in excluded_rows))

    def test_main_reports_scheduled_sources_when_delivery_validation_fails(self):
        with tempfile.TemporaryDirectory() as tmp:
            args = argparse.Namespace(
                output=Path(tmp) / "it-backend-jobs-2026-07-28.xlsx",
                max_details=200,
                sources=None,
                send_telegram=True,
                chat_id=CHAT_ID,
                token_env="TELEGRAM_BOT_TOKEN",
                memory_path=None,
                research_ledger=Path(tmp) / "researched.json",
                delivery_state=Path(tmp) / "delivery.json",
                force_send=False,
                quiet=True,
            )
            result = CollectionResult([], [], 0, 0, {})
            output = io.StringIO()
            with patch("scripts.collect_it_jobs.parse_args", return_value=args), patch(
                "scripts.collect_it_jobs.kst_now", return_value=datetime(2026, 8, 2, 9, 0)
            ), patch("scripts.collect_it_jobs.scan_official_sources", return_value=[]), patch(
                "scripts.collect_it_jobs.collect_wanted_jobs", return_value=result
            ), patch(
                "scripts.collect_it_jobs.collect_rallit_jobs", return_value=result
            ), patch("scripts.collect_it_jobs.build_workbook"), patch(
                "scripts.collect_it_jobs.validate_workbook_for_delivery", return_value=["missing workbook"]
            ), redirect_stdout(output):
                self.assertEqual(main(), 2)

        self.assertIn("sources=wanted,rallit", output.getvalue())

    def test_filters_previous_day_job_unless_due_today_or_tomorrow(self):
        previous_rows = [
            ("반복회사", "일반 백엔드 개발자", "https://example.com/normal"),
            ("오늘회사", "오늘 마감 서버 개발자", "https://example.com/today"),
            ("내일회사", "D-1 API 개발자", "https://example.com/tomorrow"),
        ]
        jobs = [
            {
                "회사명": company,
                "공고명": title,
                "링크": link,
                "마감일": deadline,
            }
            for company, title, link, deadline in [
                ("반복회사", "일반 백엔드 개발자", "https://example.com/normal", "2026-07-31"),
                ("오늘회사", "오늘 마감 서버 개발자", "https://example.com/today", "2026-07-21"),
                ("내일회사", "D-1 API 개발자", "https://example.com/tomorrow", "D-1"),
                ("신규회사", "Spring 백엔드 개발자", "https://example.com/new", "2026-07-31"),
            ]
        ]

        with tempfile.TemporaryDirectory() as tmp:
            previous_path = Path(tmp) / "it-backend-jobs-2026-07-20.xlsx"
            from openpyxl import Workbook

            previous = Workbook()
            sheet = previous.active
            sheet.title = "채용공고"
            sheet.append(["회사명", "공고명", "링크"])
            for row in previous_rows:
                sheet.append(row)
            previous.save(previous_path)

            eligible, excluded, note = filter_previous_day_duplicates(
                jobs, previous_path, "2026-07-21", "2026-07-21 09:00 KST"
            )

        self.assertEqual([job["회사명"] for job in eligible], ["오늘회사", "내일회사", "신규회사"])
        self.assertEqual(len(excluded), 1)
        self.assertEqual(excluded[0]["회사명"], "반복회사")
        self.assertIn("전일 포함 공고", excluded[0]["제외 이유"])
        self.assertIn("제외 1건", note)

    def test_keeps_jobs_when_previous_day_workbook_is_missing(self):
        jobs = [{"회사명": "신규회사", "공고명": "백엔드 개발자", "링크": "https://example.com/new"}]

        with tempfile.TemporaryDirectory() as tmp:
            eligible, excluded, note = filter_previous_day_duplicates(
                jobs,
                Path(tmp) / "it-backend-jobs-2026-07-20.xlsx",
                "2026-07-21",
                "2026-07-21 09:00 KST",
            )

        self.assertEqual(eligible, jobs)
        self.assertEqual(excluded, [])
        self.assertIn("중복 비교 미실시", note)

    def test_excludes_explicit_five_year_requirement_from_detail(self):
        search_job = {
            "id": 372829,
            "position": "백엔드 개발자(5~10년)",
            "annual_from": 5,
            "annual_to": 10,
            "company": {"name": "더스윙"},
            "due_time": "2026-07-31",
        }
        detail = {
            "job": {
                "detail": {
                    "requirements": "5년 이상의 Back-end 개발 및 운영 경험",
                    "main_tasks": "Kotlin + Spring 기반 서버 애플리케이션 개발",
                    "intro": "",
                    "preferred_points": "1년 단위 운영 경험 회고 가능자",
                }
            }
        }

        result = classify_wanted_job(search_job, detail, "2026-07-05 13:00 KST")

        self.assertIsNone(result.included)
        self.assertIn("5년 이상", result.excluded["제외 이유"])

    def test_includes_junior_backend_when_detail_matches(self):
        search_job = {
            "id": 1,
            "position": "[신입] 백엔드/서버 개발자",
            "annual_from": 0,
            "annual_to": 1,
            "company": {"name": "테스트회사"},
            "due_time": "2026-07-31",
            "address": {"full_location": "서울 강남구"},
        }
        detail = {
            "job": {
                "detail": {
                    "requirements": "Java 또는 Spring 기반 API 개발 경험",
                    "main_tasks": "백엔드 API 개발 및 DB 연동",
                    "intro": "정규직",
                    "preferred_points": "AWS 경험 우대",
                }
            }
        }

        result = classify_wanted_job(search_job, detail, "2026-07-05 13:00 KST")

        self.assertIsNotNone(result.included)
        self.assertEqual(result.included["2년차 적합도"], "상")
        self.assertEqual(result.included["백엔드 적합도"], "상")
        self.assertEqual(result.included["지원 판단"], "오늘 지원")

    def test_marks_three_year_floor_as_challenge(self):
        search_job = {
            "id": 2,
            "position": "서버 프로그래머",
            "annual_from": 3,
            "annual_to": 5,
            "company": {"name": "테스트게임"},
            "due_time": None,
        }
        detail = {
            "job": {
                "detail": {
                    "requirements": "3년 이상의 서버 개발 경험, Python 또는 Java 경험",
                    "main_tasks": "게임 서버 API 개발 및 운영",
                    "intro": "",
                    "preferred_points": "1년 단위 운영 경험 회고 가능자",
                }
            }
        }

        result = classify_wanted_job(search_job, detail, "2026-07-05 13:00 KST")

        self.assertIsNotNone(result.included)
        self.assertEqual(result.included["2년차 적합도"], "중")
        self.assertEqual(result.included["지원 판단"], "도전 지원")

    def test_excludes_non_development_title_false_positive(self):
        search_job = {
            "id": 3,
            "position": "일본향 퍼포먼스 마케터(큐텐 JP 싱글원 및 메타 셀프서버)",
            "annual_from": 1,
            "annual_to": 5,
            "company": {"name": "마케팅회사"},
            "due_time": None,
        }
        detail = {
            "job": {
                "detail": {
                    "requirements": "마케팅 캠페인 운영 경험",
                    "main_tasks": "광고 성과 분석 및 마케팅 운영",
                    "intro": "",
                    "preferred_points": "",
                }
            }
        }

        result = classify_wanted_job(search_job, detail, "2026-07-05 13:00 KST")

        self.assertIsNone(result.included)
        self.assertIn("개발 직무", result.excluded["제외 이유"])

    def test_builds_three_sheet_workbook_with_links(self):
        job = {
            "수집일": "2026-07-05",
            "우선순위": "상",
            "회사명": "테스트회사",
            "공고명": "[신입] 백엔드/서버 개발자",
            "직무분류": "Backend Engineer",
            "경력조건": "0~1년",
            "2년차 적합도": "상",
            "백엔드 적합도": "상",
            "고용형태": "미확인",
            "지역/근무형태": "서울 강남구",
            "마감일": "2026-07-31",
            "주요업무 요약": "백엔드 API 개발 및 DB 연동",
            "필수기술": "Java 또는 Spring 기반 API 개발 경험",
            "우대기술": "AWS 경험 우대",
            "지원 판단": "오늘 지원",
            "지원 메모": "2년차 백엔드에게 적합한 신입/주니어 서버 공고",
            "출처": "Wanted",
            "링크": "https://www.wanted.co.kr/wd/1",
            "확인 수준": "플랫폼 상세",
            "확인일시": "2026-07-05 13:00 KST",
        }
        excluded = {
            "회사명": "더스윙",
            "공고명": "백엔드 개발자(5~10년)",
            "링크": "https://www.wanted.co.kr/wd/372829",
            "출처": "Wanted 상세 API",
            "제외 이유": "5년 이상 요구",
            "확인일시": "2026-07-05 13:00 KST",
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "jobs.xlsx"
            build_workbook(path, "2026-07-05", [job], [excluded], ["Wanted"], "테스트")
            workbook = load_workbook(path)

        self.assertEqual(workbook.sheetnames, ["요약", "채용공고", "제외_검토"])
        self.assertEqual(workbook["채용공고"].max_row, 2)
        jobs_sheet = workbook["채용공고"]
        self.assertEqual([cell.value for cell in jobs_sheet[1]], JOB_HEADERS)
        self.assertEqual(jobs_sheet["A2"].hyperlink.target, "https://www.wanted.co.kr/wd/1")
        self.assertEqual(jobs_sheet["G2"].value, "백엔드 API 개발 및 DB 연동")
        self.assertEqual(jobs_sheet.freeze_panes, "A2")
        self.assertTrue(jobs_sheet.auto_filter.ref)
        self.assertFalse(jobs_sheet.sheet_view.showGridLines)
        self.assertGreater(jobs_sheet.column_dimensions["G"].width, 30)
        summary_values = [cell.value for row in workbook["요약"].iter_rows() for cell in row]
        self.assertNotIn("오늘 우선 확인 TOP 5: 회사명", summary_values)
        self.assertEqual(workbook["제외_검토"].max_row, 2)

    def test_summary_message_uses_counts_without_top_five(self):
        jobs = [
            {
                "회사명": "테스트회사",
                "공고명": "[신입] 백엔드/서버 개발자",
                "지원 메모": "신입 서버 API라 바로 확인할 가치가 높음",
                "2년차 적합도": "상",
                "마감일": "2026-07-10",
            }
        ]

        message = build_summary_message("2026-07-05", 3, jobs)

        self.assertIn("총 확인: 3개", message)
        self.assertIn("엑셀 포함: 1개", message)
        self.assertIn("적합도 상: 1개", message)
        self.assertNotIn("TOP 5", message)
        self.assertNotIn("테스트회사 - [신입] 백엔드/서버 개발자", message)

    def test_summary_message_counts_today_deadline_as_urgent(self):
        jobs = [
            {
                "회사명": "오늘회사",
                "공고명": "오늘 마감 백엔드 개발자",
                "지원 메모": "오늘 지원 필요",
                "2년차 적합도": "상",
                "마감일": "D-0",
            }
        ]

        message = build_summary_message("2026-07-21", 1, jobs)

        self.assertIn("마감 7일 이내: 1개", message)

    def test_includes_rallit_beginner_backend_detail(self):
        list_item = {
            "id": 4252,
            "title": "[채용전환형] Backend Engineer",
            "companyName": "랠릿회사",
            "endedAt": "2026-07-19",
            "jobLevels": ["BEGINNER"],
            "url": "https://www.rallit.com/positions/4252",
            "addressRegion": "GANGNAM",
            "jobSkillKeywords": ["Python", "Django", "PostgreSQL", "AWS"],
            "status": {"code": "HIRING", "name": "모집 중"},
        }
        detail = {
            "id": 4252,
            "title": "[채용전환형] Backend Engineer",
            "companyName": "랠릿회사",
            "endedAt": "2026-07-19",
            "jobLevels": ["BEGINNER"],
            "jobs": [{"code": "BACKEND_DEVELOPER", "name": "백엔드/서버 개발자"}],
            "responsibilities": "<p>백엔드 API 개발 및 PostgreSQL 기반 데이터 모델링</p>",
            "basicQualifications": "<p>Django 또는 Spring 기반 웹 서비스 개발 경험</p>",
            "preferredQualifications": "<p>AWS 운영 경험 우대</p>",
            "jobSkillKeywords": ["Python", "Django", "PostgreSQL", "AWS"],
            "addressMain": "서울 강남구 테헤란로",
            "status": {"code": "HIRING", "name": "모집 중"},
        }

        result = classify_rallit_position(list_item, detail, "2026-07-05 13:00 KST")

        self.assertIsNotNone(result.included)
        self.assertEqual(result.included["출처"], "Rallit")
        self.assertEqual(result.included["2년차 적합도"], "상")
        self.assertEqual(result.included["지원 판단"], "오늘 지원")

    def test_select_jobs_for_workbook_keeps_source_mix(self):
        wanted_jobs = [
            {"회사명": f"원티드{i}", "출처": "Wanted", "우선순위": "상", "2년차 적합도": "상", "마감일": "2026-07-31"}
            for i in range(30)
        ]
        rallit_jobs = [
            {"회사명": "랠릿1", "출처": "Rallit", "우선순위": "상", "2년차 적합도": "상", "마감일": "2026-07-31"}
        ]

        selected = select_jobs_for_workbook(wanted_jobs + rallit_jobs, "2026-07-05", limit=25)

        self.assertEqual(len(selected), 25)
        self.assertIn("Rallit", {job["출처"] for job in selected})

    def test_select_jobs_for_workbook_defaults_to_two_hundred(self):
        jobs = [
            {
                "회사명": f"회사{i}",
                "공고명": f"백엔드 개발자{i}",
                "출처": "Wanted",
                "우선순위": "상",
                "2년차 적합도": "상",
                "마감일": "2026-07-31",
            }
            for i in range(220)
        ]

        self.assertEqual(len(select_jobs_for_workbook(jobs, "2026-07-28")), 200)

    def test_select_jobs_for_workbook_deduplicates_same_company_and_title(self):
        duplicate_a = {
            "회사명": "중복회사",
            "공고명": "백엔드 개발자",
            "출처": "JobKorea",
            "우선순위": "상",
            "2년차 적합도": "상",
            "마감일": "2026-07-31",
            "링크": "https://www.jobkorea.co.kr/Recruit/GI_Read/1?stext=백엔드",
        }
        duplicate_b = {
            "회사명": "중복회사",
            "공고명": "백엔드 개발자",
            "출처": "JobKorea",
            "우선순위": "상",
            "2년차 적합도": "상",
            "마감일": "2026-07-31",
            "링크": "https://www.jobkorea.co.kr/Recruit/GI_Read/1?stext=서버",
        }
        distinct = {
            "회사명": "다른회사",
            "공고명": "서버 개발자",
            "출처": "JobKorea",
            "우선순위": "상",
            "2년차 적합도": "상",
            "마감일": "2026-07-31",
            "링크": "https://www.jobkorea.co.kr/Recruit/GI_Read/2",
        }

        selected = select_jobs_for_workbook([duplicate_a, duplicate_b, distinct], "2026-07-05", limit=3)

        pairs = [(job["회사명"], job["공고명"]) for job in selected]
        self.assertEqual(pairs.count(("중복회사", "백엔드 개발자")), 1)
        self.assertEqual(len(selected), 2)

    def test_includes_saramin_detail_from_meta_description(self):
        html = """
        <html><head>
        <title>[(주)인터엑스] 백엔드(Backend)개발자(D-6) - 사람인</title>
        <meta name="description" content="(주)인터엑스, 백엔드(Backend)개발자, 경력:신입/경력, 학력:대학졸업(2,3년)이상, 회사내규에 따름, 마감일:2026-07-11, 홈페이지:www.interxlab.com">
        </head><body>백엔드 API 개발 Java Spring</body></html>
        """

        result = classify_saramin_detail(
            "https://www.saramin.co.kr/zf_user/jobs/relay/view?rec_idx=1",
            html,
            "2026-07-05 13:00 KST",
        )

        self.assertIsNotNone(result.included)
        self.assertEqual(result.included["출처"], "Saramin")
        self.assertEqual(result.included["경력조건"], "신입/경력")

    def test_includes_jobkorea_detail_from_body_text(self):
        html = """
        <html><head>
        <title>(주)우아한형제들 채용 - 각 부문별 경력/신입 인재영입 | 잡코리아</title>
        <meta name="description" content="경력 : 신입·경력 , 학력 : 학력무관, 급여 : 회사 내규에 따름, 마감일 : 상시채용">
        </head><body>
        <h2>(주)우아한형제들</h2>
        모집요강 모집분야 백엔드 서버 개발자 지원자격 경력 신입·경력 학력 학력무관
        담당업무 Java Spring 기반 API 개발 마감일 상시채용
        </body></html>
        """

        result = classify_jobkorea_detail(
            "https://www.jobkorea.co.kr/Recruit/GI_Read/1",
            html,
            "2026-07-05 13:00 KST",
        )

        self.assertIsNotNone(result.included)
        self.assertEqual(result.included["출처"], "JobKorea")
        self.assertEqual(result.included["지원 판단"], "오늘 지원")

    def test_excludes_jobkorea_ten_year_requirement(self):
        html = """
        <html><head>
        <title>GS리테일 채용 - AX본부 인프라부문 물류AX 담당 경력사원 채용 | 잡코리아</title>
        <meta name="description" content="경력 : 경력 10년이상, 마감일 : 2026-07-31">
        </head><body>지원자격 경력 10년이상 물류 AX 인프라 운영</body></html>
        """

        result = classify_jobkorea_detail("https://www.jobkorea.co.kr/Recruit/GI_Read/2", html, "2026-07-05 13:00 KST")

        self.assertIsNone(result.included)
        self.assertIn("5년 이상", result.excluded["제외 이유"])

    def test_excludes_broad_jobkorea_non_backend_collection_post(self):
        html = """
        <html><head>
        <title>NC 채용 - 엔씨소프트 전 부문 수시 채용 | 잡코리아</title>
        <meta name="description" content="경력 : 신입·경력 , 마감일 : 상시채용">
        </head><body>게임 사업, 경영지원, 개발 등 전 부문 채용 검색어 백엔드</body></html>
        """

        result = classify_jobkorea_detail("https://www.jobkorea.co.kr/Recruit/GI_Read/3", html, "2026-07-05 13:00 KST")

        self.assertIsNone(result.included)
        self.assertIn("백엔드 맞춤성이 낮음", result.excluded["제외 이유"])

    def test_excludes_jobkorea_search_keyword_only_false_positive(self):
        html = """
        <html><head>
        <title>한국산업은행 채용 - 한국산업은행 2026년 2차 청년인턴 채용 공고 | 잡코리아</title>
        <meta name="description" content="경력 : 신입·경력 , 학력 : 학력무관, 마감일 : 2026-07-31">
        </head><body>
        금융권 청년인턴 채용 공고입니다. 검색어 백엔드 결과 목록으로 돌아가기
        </body></html>
        """

        result = classify_jobkorea_detail("https://www.jobkorea.co.kr/Recruit/GI_Read/4", html, "2026-07-05 13:00 KST")

        self.assertIsNone(result.included)
        self.assertIn("백엔드 맞춤성이 낮음", result.excluded["제외 이유"])

    def test_excludes_jobkorea_restaurant_hall_false_positive(self):
        html = """
        <html><head>
        <title>까사 이태리아 채용 - 종로/서촌 이탈리안 레스토랑 까사 이태리아 ( 홀 신입/경력 ) 채용 | 잡코리아</title>
        <meta name="description" content="경력 : 신입·경력 , 마감일 : 2026-07-31">
        </head><body>
        담당업무 홀 서비스, 고객 응대, 레스토랑 매장 운영
        검색어 백엔드 채용정보 개발자 공고
        </body></html>
        """

        result = classify_jobkorea_detail("https://www.jobkorea.co.kr/Recruit/GI_Read/6", html, "2026-07-05 13:00 KST")

        self.assertIsNone(result.included)
        self.assertIn("비개발", result.excluded["제외 이유"])

    def test_includes_it_relevant_youth_intern_posting(self):
        html = """
        <html><head>
        <title>한국산업은행 채용 - 디지털·IT 청년인턴 채용 공고 | 잡코리아</title>
        <meta name="description" content="경력 : 신입·경력 , 학력 : 학력무관, 마감일 : 2026-07-31">
        </head><body>
        모집분야 디지털 IT 청년인턴 담당업무 데이터 분석, 정보시스템 운영, API 연계 업무 지원
        지원자격 신입·경력 학력무관
        </body></html>
        """

        result = classify_jobkorea_detail("https://www.jobkorea.co.kr/Recruit/GI_Read/5", html, "2026-07-05 13:00 KST")

        self.assertIsNotNone(result.included)
        self.assertEqual(result.included["직무분류"], "Youth Intern / IT")
        self.assertEqual(result.included["지원 판단"], "오늘 지원")

    def test_excludes_non_it_youth_intern_false_positive(self):
        html = """
        <html><head>
        <title>[(주)하이테커] K-디자인 청년 인턴 모집(콘텐츠/마케팅/영상)(D-8) - 사람인</title>
        <meta name="description" content="경력 : 신입, 마감일 : 2026-07-13">
        </head><body>
        담당업무 콘텐츠 제작, 마케팅 운영, 영상 편집, 디자인 보조
        site navigation position list
        </body></html>
        """

        result = classify_saramin_detail(
            "https://www.saramin.co.kr/zf_user/jobs/relay/view?rec_idx=2",
            html,
            "2026-07-05 13:00 KST",
        )

        self.assertIsNone(result.included)
        self.assertIn("백엔드 맞춤성이 낮음", result.excluded["제외 이유"])

    def test_excludes_rallit_frontend_only_even_if_beginner(self):
        detail = {
            "id": 5,
            "title": "프론트엔드 개발자 (React.JS_주니어)",
            "companyName": "프론트회사",
            "jobLevels": ["BEGINNER"],
            "jobs": [{"code": "FRONTEND_DEVELOPER", "name": "프론트엔드 개발자"}],
            "responsibilities": "<p>React 화면 개발</p>",
            "basicQualifications": "<p>JavaScript React 경험</p>",
            "preferredQualifications": "",
            "jobSkillKeywords": ["React", "JavaScript"],
            "status": {"code": "HIRING", "name": "모집 중"},
        }

        result = classify_rallit_position(detail, detail, "2026-07-05 13:00 KST")

        self.assertIsNone(result.included)
        self.assertIn("프론트/모바일", result.excluded["제외 이유"])

    def test_excludes_rallit_korean_android_only_even_if_beginner(self):
        detail = {
            "id": 7,
            "title": "안드로이드 개발자",
            "companyName": "모바일회사",
            "jobLevels": ["BEGINNER"],
            "jobs": [{"code": "ANDROID_DEVELOPER", "name": "안드로이드 개발자"}],
            "responsibilities": "<p>안드로이드 앱 개발, 유지보수</p>",
            "basicQualifications": "<p>Java 또는 Kotlin 개발 언어에 능숙한 분</p>",
            "preferredQualifications": "",
            "jobSkillKeywords": ["Android", "Kotlin", "Java"],
            "status": {"code": "HIRING", "name": "모집 중"},
        }

        result = classify_rallit_position(detail, detail, "2026-07-08 09:30 KST")

        self.assertIsNone(result.included)
        self.assertIn("프론트/모바일", result.excluded["제외 이유"])

    def test_excludes_rallit_five_to_twelve_year_range(self):
        detail = {
            "id": 6,
            "title": "Backend Kotlin Developer (5-12년)",
            "companyName": "시니어회사",
            "jobLevels": ["BEGINNER"],
            "jobs": [{"code": "BACKEND_DEVELOPER", "name": "백엔드/서버 개발자"}],
            "responsibilities": "<p>Kotlin 백엔드 API 개발</p>",
            "basicQualifications": "<p>Backend Kotlin Developer 5-12년 경력</p>",
            "preferredQualifications": "",
            "jobSkillKeywords": ["Kotlin", "Spring"],
            "status": {"code": "HIRING", "name": "모집 중"},
        }

        result = classify_rallit_position(detail, detail, "2026-07-05 13:00 KST")

        self.assertIsNone(result.included)
        self.assertIn("5년 이상", result.excluded["제외 이유"])

    def test_parse_source_names_supports_all_and_subset(self):
        self.assertEqual(
            parse_source_names("all"),
            ["wanted", "rallit", "saramin", "jobkorea", *FALLBACK_SOURCES, "career"],
        )
        self.assertEqual(parse_source_names("wanted,jobkorea"), ["wanted", "jobkorea"])

    def test_parse_source_names_rejects_unknown_source(self):
        with self.assertRaises(ValueError):
            parse_source_names("wanted,unknown")

    def test_parse_args_accepts_sources_and_quiet_mode(self):
        args = parse_args(["--sources", "wanted,jobkorea", "--quiet"])

        self.assertEqual(args.sources, ["wanted", "jobkorea"])
        self.assertTrue(args.quiet)

    def test_send_telegram_rejects_non_final_attachment_name_before_network(self):
        ok, status = send_telegram(
            "dummy-token",
            "6907667924",
            "message",
            Path("/tmp/it-backend-jobs-2026-07-05-candidate-2.xlsx"),
        )

        self.assertFalse(ok)
        self.assertIn("non-final", status)

    def test_delivery_state_blocks_second_send_for_same_date_and_chat(self):
        with tempfile.TemporaryDirectory() as tmp:
            state_path = Path(tmp) / "delivery.json"
            key = delivery_key("2026-07-05", "6907667924")

            self.assertFalse(delivery_already_sent(state_path, key))
            record_delivery(state_path, key, "/tmp/it-backend-jobs-2026-07-05.xlsx", "2026-07-05 09:00 KST")

            self.assertTrue(delivery_already_sent(state_path, key))

    def test_platform_search_terms_include_backend_and_youth_interns(self):
        terms = platform_search_terms()

        self.assertIn("백엔드", terms)
        self.assertIn("디지털 청년인턴", terms)
        self.assertLess(terms.index("디지털 청년인턴"), terms.index("Java Spring"))
        self.assertEqual(len(terms), len(set(terms)))

    def test_merge_platform_links_caps_each_keyword_bucket(self):
        links = merge_platform_links(
            [
                ["https://example.com/backend-1", "https://example.com/backend-2", "https://example.com/backend-3"],
                ["https://example.com/youth-1", "https://example.com/youth-2"],
            ],
            max_details=4,
        )

        self.assertEqual(
            links,
            [
                "https://example.com/backend-1",
                "https://example.com/backend-2",
                "https://example.com/youth-1",
                "https://example.com/youth-2",
            ],
        )

    def test_validate_workbook_rejects_non_it_youth_intern_rows(self):
        job = {
            "수집일": "2026-07-05",
            "우선순위": "중",
            "회사명": "(주)하이테커",
            "공고명": "K-디자인 청년 인턴 모집(콘텐츠/마케팅/영상)",
            "직무분류": "Youth Intern / IT",
            "경력조건": "신입",
            "2년차 적합도": "중",
            "백엔드 적합도": "중",
            "고용형태": "미확인",
            "지역/근무형태": "미확인",
            "마감일": "2026-07-14",
            "주요업무 요약": "콘텐츠 제작, 마케팅 운영, 영상 편집",
            "필수기술": "경력: 신입",
            "우대기술": "미확인",
            "지원 판단": "저장",
            "지원 메모": "청년인턴",
            "출처": "Saramin",
            "링크": "https://example.com/job",
            "확인 수준": "플랫폼 상세",
            "확인일시": "2026-07-05 13:00 KST",
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "jobs.xlsx"
            build_workbook(path, "2026-07-05", [job], [], ["Saramin"], "테스트")
            errors = validate_workbook_for_delivery(path, [job])

        self.assertTrue(any("비IT 청년인턴" in error for error in errors))

    def test_validate_workbook_accepts_valid_youth_intern_rows(self):
        job = {
            "수집일": "2026-07-05",
            "우선순위": "중",
            "회사명": "국토교통부",
            "공고명": "청년인턴 채용 공고(전산)",
            "직무분류": "Youth Intern / IT",
            "경력조건": "신입",
            "2년차 적합도": "중",
            "백엔드 적합도": "중",
            "고용형태": "미확인",
            "지역/근무형태": "미확인",
            "마감일": "2026-07-14",
            "주요업무 요약": "전산 시스템 운영 지원",
            "필수기술": "전산 관련 업무",
            "우대기술": "미확인",
            "지원 판단": "저장",
            "지원 메모": "IT 청년인턴",
            "출처": "Saramin",
            "링크": "https://example.com/job",
            "확인 수준": "플랫폼 상세",
            "확인일시": "2026-07-05 13:00 KST",
        }

        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "jobs.xlsx"
            build_workbook(path, "2026-07-05", [job], [], ["Saramin"], "테스트")
            errors = validate_workbook_for_delivery(path, [job])

        self.assertEqual(errors, [])


if __name__ == "__main__":
    unittest.main()
