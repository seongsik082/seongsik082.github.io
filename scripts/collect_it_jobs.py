#!/usr/bin/env python3
"""Collect backend-friendly IT jobs, write XLSX, and optionally send Telegram."""

from __future__ import annotations

import argparse
import html
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover - Python 3.8 fallback only.
    ZoneInfo = None


CHAT_ID = "6907667924"
DEFAULT_DELIVERY_STATE = Path("/tmp/it-backend-jobs-delivery-state.json")
DEFAULT_RESEARCH_LEDGER = Path("/Users/kss/.codex/automations/it/researched-jobs.json")
WANTED_SEARCH_URL = "https://www.wanted.co.kr/api/v4/jobs"
WANTED_DETAIL_URL = "https://www.wanted.co.kr/api/v4/jobs/{job_id}"
RALLIT_LIST_URL = "https://www.rallit.com/client/api/v1/position"
RALLIT_DETAIL_URL = "https://www.rallit.com/client/api/v1/position/{position_id}"
SARAMIN_SEARCH_URL = "https://www.saramin.co.kr/zf_user/search/recruit"
JOBKOREA_SEARCH_URL = "https://www.jobkorea.co.kr/Search/"
JUMPIT_LIST_URL = "https://jumpit-api.saramin.co.kr/api/positions"
JUMPIT_DETAIL_URL = "https://jumpit-api.saramin.co.kr/api/position/{position_id}"
CATCH_LIST_URL = "https://m.catch.co.kr/api/v1.0/recruit/information/getRecruitList"
INCRUIT_SEARCH_URL = "https://job.incruit.com/jobdb_list/searchjob.asp"
WORKNET_SEARCH_URL = "https://www.work.go.kr/empInfo/empInfoSrch/list/dtlEmpSrchList.do"
STARTING_LIST_URL = "https://starting.kr/api/jdlist/postings"
STARTING_DETAIL_URL = "https://app.starting.kr/matching/{public_number}"

KEYWORDS = [
    "백엔드",
    "서버",
    "Java Spring",
    "API",
    "플랫폼 엔지니어",
    "DevOps",
    "데이터 엔지니어",
]

FALLBACK_SOURCES = ("jumpit", "catch", "incruit", "worknet", "starting")
AVAILABLE_SOURCES = ("wanted", "rallit", "saramin", "jobkorea", *FALLBACK_SOURCES)
SCHEDULE_ANCHOR_DATE = date(2026, 7, 30)
TARGET_INCLUDED_JOBS = 200
YOUTH_INTERN_SEARCH_TERMS = ["디지털 청년인턴", "IT 청년인턴", "전산 청년인턴"]
PLATFORM_PRIORITY_SEARCH_TERMS = ["백엔드", "서버"] + YOUTH_INTERN_SEARCH_TERMS

OFFICIAL_SOURCES = [
    ("당근", "https://jobs.daangn.com/"),
    ("카카오", "https://careers.kakao.com/jobs"),
    ("토스", "https://toss.im/career/jobs"),
    ("우아한형제들", "https://career.woowahan.com/"),
    ("NHN", "https://careers.nhn.com/"),
    ("NAVER", "https://career.navercorp.com/"),
]

JOB_HEADERS = ["링크", "회사명", "공고명", "경력조건", "지역/근무형태", "마감일", "주요업무요약", "출처"]
JOB_SHEET_FIELDS = {
    "링크": "링크",
    "회사명": "회사명",
    "공고명": "공고명",
    "경력조건": "경력조건",
    "지역/근무형태": "지역/근무형태",
    "마감일": "마감일",
    "주요업무요약": "주요업무 요약",
    "출처": "출처",
}

EXCLUDED_HEADERS = ["회사명", "공고명", "링크", "출처", "제외 이유", "확인일시"]

BACKEND_TERMS = [
    "백엔드",
    "서버",
    "backend",
    "server",
    "spring",
    "java",
    "kotlin",
    "api",
    "platform",
    "플랫폼",
    "cloud",
    "클라우드",
    "devops",
    "infra",
    "인프라",
    "database",
    "db",
    "데이터",
    "data engineer",
]

JUNIOR_TERMS = ["신입", "인턴", "주니어", "junior", "intern", "경력무관", "0년", "1년", "2년"]
TITLE_EXCLUDE_TERMS = ["팀장", "매니저", "manager", "lead", "리드", "시니어", "senior", "tech lead", "테크 리드"]
NON_BACKEND_CENTERED_TERMS = [
    "frontend",
    "프론트엔드",
    "ios",
    "android",
    "안드로이드",
    "아이폰",
    "모바일",
    "ai research",
    "research scientist",
]
NON_DEVELOPMENT_TITLE_TERMS = [
    "마케터",
    "퍼포먼스 마케터",
    "기술지원",
    "하드웨어 기술지원",
    "fae",
    "기구 설계",
    "영업",
    "기획자",
    "서비스 기획",
    "레스토랑",
    "홀 신입",
    "홀 경력",
    "홀 서비스",
    "주방",
    "서빙",
    "매장 운영",
]
BROAD_COLLECTION_TERMS = ["전 부문", "전직군", "각 부문", "수시 채용", "공채", "신입공채"]
YOUTH_INTERN_TERMS = ["청년인턴", "청년 인턴", "체험형 인턴"]
IT_INTERN_TERMS = ["디지털", "IT", "SW", "소프트웨어", "개발", "데이터", "전산", "정보시스템", "보안", "인프라", "클라우드", "API"]
NON_IT_YOUTH_INTERN_TITLE_TERMS = ["콘텐츠", "마케팅", "영상", "디자인", "영업", "사무", "행정", "홍보", "회계"]


@dataclass
class ClassifiedResult:
    included: Optional[Dict[str, object]]
    excluded: Optional[Dict[str, object]]


@dataclass
class CollectionResult:
    included: List[Dict[str, object]]
    excluded: List[Dict[str, object]]
    checked: int
    skipped_researched: int
    research_records: Dict[str, Dict[str, str]]


def kst_now() -> datetime:
    if ZoneInfo is None:
        return datetime.now()
    return datetime.now(ZoneInfo("Asia/Seoul"))


def clean_text(value: object, limit: int = 500) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text[:limit]


def contains_any(text: str, terms: Iterable[str]) -> bool:
    lower = text.lower()
    return any(term.lower() in lower for term in terms)


def parse_source_names(value: str) -> List[str]:
    if not value or value.strip().lower() == "all":
        return list(AVAILABLE_SOURCES)
    requested = {part.strip().lower() for part in value.split(",") if part.strip()}
    invalid = sorted(requested.difference(AVAILABLE_SOURCES))
    if invalid:
        raise ValueError(f"unknown sources: {', '.join(invalid)}; choose from {', '.join(AVAILABLE_SOURCES)} or all")
    return [source for source in AVAILABLE_SOURCES if source in requested]


def scheduled_sources(collect_date: str) -> List[str]:
    cycle_day = (date.fromisoformat(collect_date) - SCHEDULE_ANCHOR_DATE).days % 4
    if cycle_day < 3:
        return ["saramin", "jobkorea"]
    return ["wanted", "rallit"]


def normalize_research_url(url: str) -> str:
    parsed = urlsplit(clean_text(url))
    ignored_query_keys = {"stext", "keyword", "searchword"}
    query = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=True)
        if key.lower() not in ignored_query_keys and not key.lower().startswith("utm_")
    ]
    return urlunsplit(
        (
            parsed.scheme.lower(),
            parsed.netloc.lower(),
            parsed.path.rstrip("/"),
            urlencode(query),
            "",
        )
    )


def load_research_ledger(path: Path) -> Dict[str, Dict[str, str]]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(data, dict):
        return {}
    return {
        normalized: {str(key): str(value) for key, value in record.items()}
        for url, record in data.items()
        for normalized in [normalize_research_url(str(url))]
        if isinstance(record, dict) and normalized
    }


def save_research_ledger(path: Path, records: Dict[str, Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(records, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def record_researched_url(
    records: Dict[str, Dict[str, str]], url: str, source: str, outcome: str, checked_at: str
) -> None:
    normalized = normalize_research_url(url)
    if not normalized:
        return
    records[normalized] = {"source": source, "outcome": outcome, "checked_at": checked_at}


def was_researched(url: str, researched_urls: Set[str]) -> bool:
    return normalize_research_url(url) in researched_urls


def append_classification(
    collection: CollectionResult,
    result: ClassifiedResult,
    url: str,
    source: str,
    checked_at: str,
) -> None:
    if result.included:
        collection.included.append(result.included)
        record_researched_url(collection.research_records, url, source, "included", checked_at)
    elif result.excluded:
        collection.excluded.append(result.excluded)
        record_researched_url(collection.research_records, url, source, "excluded", checked_at)


def platform_search_terms() -> List[str]:
    return list(dict.fromkeys(PLATFORM_PRIORITY_SEARCH_TERMS + KEYWORDS))


def merge_platform_links(search_results: Sequence[Sequence[str]], max_details: int) -> List[str]:
    if max_details <= 0:
        return []
    bucket_count = max(1, min(len(search_results), 4))
    per_bucket = max(1, max_details // bucket_count)
    links: List[str] = []
    seen = set()
    for bucket in search_results:
        bucket_added = 0
        for url in bucket:
            if url in seen:
                continue
            seen.add(url)
            links.append(url)
            bucket_added += 1
            if len(links) >= max_details or bucket_added >= per_bucket:
                break
        if len(links) >= max_details:
            break
    return links


def explicit_five_plus(text: str) -> bool:
    patterns = [
        r"5\s*년\s*이상",
        r"5\s*\+\s*년",
        r"5\s*years?\s*\+",
        r"over\s*5\s*years?",
        r"more\s*than\s*5\s*years?",
    ]
    return any(re.search(pattern, text, re.IGNORECASE) for pattern in patterns)


def explicit_year_floor_at_least(text: str, minimum: int) -> bool:
    for match in re.finditer(r"(\d+)\s*[-~]\s*(\d+)\s*년", text):
        try:
            if int(match.group(1)) >= minimum:
                return True
        except ValueError:
            continue
    for match in re.finditer(r"(\d+)\s*년\s*(?:이상|~)", text):
        try:
            if int(match.group(1)) >= minimum:
                return True
        except ValueError:
            continue
    for match in re.finditer(r"(?:경력|개발\s*경력)\s*(\d+)\s*년차", text):
        try:
            if int(match.group(1)) >= minimum:
                return True
        except ValueError:
            continue
    return False


def title_has_backend_signal(title: str) -> bool:
    return contains_any(title, ["백엔드", "서버", "backend", "server", "api", "devops", "플랫폼", "platform", "data engineer", "데이터 엔지니어", "인프라"])


def title_has_non_backend_focus(title: str) -> bool:
    return contains_any(title, NON_BACKEND_CENTERED_TERMS) and not title_has_backend_signal(title) and not contains_any(title, ["풀스택", "fullstack", "full-stack"])


def structured_backend_signal(text: str) -> bool:
    return bool(re.search(r"(모집분야|담당업무|지원자격|주요업무|자격요건).{0,120}(백엔드|서버|API|Spring|Java|Kotlin|Node)", text, re.IGNORECASE))


def platform_backend_signal(title: str, description: str, visible_text: str) -> bool:
    return title_has_backend_signal(title) or contains_any(description, ["백엔드", "서버", "backend", "server", "api", "spring", "java", "kotlin", "devops"]) or structured_backend_signal(visible_text)


def contains_it_intern_term(text: str) -> bool:
    plain_terms = [term for term in IT_INTERN_TERMS if term not in {"IT", "SW", "API"}]
    return contains_any(text, plain_terms) or bool(re.search(r"(?<![A-Za-z])(IT|SW|API)(?![A-Za-z])", text, re.IGNORECASE))


def has_nearby_youth_it_signal(text: str) -> bool:
    youth = r"(청년\s*인턴|청년인턴|체험형\s*인턴)"
    it = r"(디지털|전산|정보시스템|소프트웨어|개발|데이터|보안|인프라|클라우드|API|IT|SW)"
    return bool(
        re.search(rf"{youth}.{{0,160}}{it}|{it}.{{0,160}}{youth}", text, re.IGNORECASE)
    )


def youth_intern_it_signal(title: str, description: str, visible_text: str) -> bool:
    combined = " ".join([title, description, visible_text])
    if not contains_any(combined, YOUTH_INTERN_TERMS):
        return False
    decisive_text = " ".join([title, description])
    if contains_it_intern_term(decisive_text):
        return True
    if contains_any(title, NON_IT_YOUTH_INTERN_TITLE_TERMS):
        return False
    return has_nearby_youth_it_signal(visible_text)


def explicit_four_plus(text: str) -> bool:
    return bool(re.search(r"4\s*년\s*이상|4\s*\+\s*년", text, re.IGNORECASE))


def summarize_text(text: str, fallback: str = "미확인", limit: int = 180) -> str:
    text = clean_text(text, limit=1000)
    if not text:
        return fallback
    pieces = re.split(r"[•\n\r]+|(?<=다\.)\s+", text)
    summary = " / ".join(clean_text(piece, 100) for piece in pieces if clean_text(piece, 100))
    return summary[:limit] if summary else text[:limit]


def html_to_text(value: object) -> str:
    if not value:
        return ""
    return clean_text(BeautifulSoup(str(value), "lxml").get_text(" "))


def meta_content(soup: BeautifulSoup, selector: str) -> str:
    tag = soup.select_one(selector)
    return clean_text(tag.get("content")) if tag and tag.get("content") else ""


def extract_title_company_from_platform_title(title: str, source_suffix: str) -> Tuple[str, str]:
    title = title.replace(source_suffix, "").strip()
    bracket = re.match(r"\[(?P<company>[^\]]+)\]\s*(?P<job>.+)", title)
    if bracket:
        return clean_text(bracket.group("company")), clean_text(bracket.group("job"))
    if " 채용 - " in title:
        company, job = title.split(" 채용 - ", 1)
        return clean_text(company), clean_text(job.replace("| 잡코리아", ""))
    return "미확인", clean_text(title)


def extract_meta_field(description: str, label: str) -> str:
    pattern = rf"{re.escape(label)}\s*[:：]\s*([^,|]+)"
    match = re.search(pattern, description)
    return clean_text(match.group(1)) if match else "미확인"


def classify_html_platform_detail(
    source: str,
    source_label: str,
    url: str,
    html_text: str,
    checked_at: str,
    collect_date: str,
) -> ClassifiedResult:
    soup = BeautifulSoup(html_text, "lxml")
    page_title = clean_text(soup.title.get_text(" ", strip=True) if soup.title else "")
    description = meta_content(soup, 'meta[name="description"]')
    visible_text = clean_text(soup.get_text(" ", strip=True), 3000)
    company, title = extract_title_company_from_platform_title(page_title, f"- {source_label}")
    if source == "JobKorea":
        company_tag = soup.select_one("h2")
        if company_tag and company_tag.get_text(strip=True):
            company = clean_text(company_tag.get_text(" ", strip=True))
    career = extract_meta_field(description, "경력")
    deadline = extract_meta_field(description, "마감일")
    all_text = " ".join([page_title, description, visible_text])
    youth_it = youth_intern_it_signal(title, description, visible_text)

    def excluded(reason: str) -> ClassifiedResult:
        return ClassifiedResult(
            included=None,
            excluded={
                "회사명": company,
                "공고명": title,
                "링크": url,
                "출처": f"{source} 상세 HTML",
                "제외 이유": reason,
                "확인일시": checked_at,
            },
        )

    if contains_any(title, TITLE_EXCLUDE_TERMS) or contains_any(title, NON_DEVELOPMENT_TITLE_TERMS):
        return excluded("제목상 리드/비개발 직무 성격이 강함")
    if explicit_five_plus(all_text) or explicit_year_floor_at_least(all_text, 5):
        return excluded("상세 본문 또는 메타에서 5년 이상 요구 확인")
    if contains_any(title, BROAD_COLLECTION_TERMS) and not title_has_backend_signal(title) and not structured_backend_signal(visible_text):
        return excluded("전 부문/공채형 공고로 2년차 백엔드 맞춤성이 낮음")
    if title_has_non_backend_focus(title):
        return excluded("프론트/모바일 중심으로 백엔드 전환 가치가 불명확")
    if source in {"Saramin", "JobKorea"} and not platform_backend_signal(title, description, visible_text) and not youth_it:
        return excluded("상세 본문/메타에서 백엔드 맞춤성이 낮음")
    if source not in {"Saramin", "JobKorea"} and not contains_any(all_text, BACKEND_TERMS):
        return excluded("상세 본문/메타에서 백엔드/API/서버/플랫폼 연관성이 부족")
    if not description and len(visible_text) < 200:
        return excluded("상세 본문 확인 불충분")

    junior = contains_any(" ".join([career, all_text]), ["신입", "경력무관", "주니어", "1년", "2년"])
    challenge = re.search(r"3\s*년\s*이상", all_text) or explicit_four_plus(all_text)
    if challenge and not junior:
        priority, fit, decision = "중", "중", "도전 지원"
        memo = f"{source_label} 상세에서 3년 이상 성격이 보여 도전 지원 후보"
    elif youth_it and not platform_backend_signal(title, description, visible_text):
        priority, fit, decision = "중", "중", "저장"
        memo = f"{source_label} 상세에서 IT/디지털 청년인턴으로 백엔드 인접 경험 가능성 확인"
    else:
        priority, fit, decision = "상", "상", "오늘 지원"
        memo = f"{source_label} 상세에서 신입/주니어 또는 2년차 접근 가능한 백엔드 공고 확인"

    job_class = "Youth Intern / IT" if youth_it else "Backend Engineer"

    included = {
        "수집일": collect_date,
        "우선순위": priority,
        "회사명": company,
        "공고명": title,
        "직무분류": job_class,
        "경력조건": career,
        "2년차 적합도": fit,
        "백엔드 적합도": "상" if contains_any(all_text, ["spring", "java", "kotlin", "api", "server", "서버", "백엔드"]) else "중",
        "고용형태": "미확인",
        "지역/근무형태": "미확인",
        "마감일": deadline,
        "주요업무 요약": summarize_text(visible_text),
        "필수기술": summarize_text(description or visible_text),
        "우대기술": "미확인",
        "지원 판단": decision,
        "지원 메모": memo,
        "출처": source,
        "링크": url,
        "확인 수준": "플랫폼 상세",
        "확인일시": checked_at,
    }
    return ClassifiedResult(included=included, excluded=None)


def classify_saramin_detail(url: str, html_text: str, checked_at: str, collect_date: str = "2026-07-05") -> ClassifiedResult:
    return classify_html_platform_detail("Saramin", "사람인", url, html_text, checked_at, collect_date)


def classify_jobkorea_detail(url: str, html_text: str, checked_at: str, collect_date: str = "2026-07-05") -> ClassifiedResult:
    return classify_html_platform_detail("JobKorea", "잡코리아", url, html_text, checked_at, collect_date)


def career_range(job: Dict[str, object]) -> str:
    annual_from = job.get("annual_from")
    annual_to = job.get("annual_to")
    if annual_from is None and annual_to is None:
        return "미확인"
    return f"{annual_from if annual_from is not None else '미확인'}~{annual_to if annual_to is not None else '미확인'}년"


def location_text(job: Dict[str, object]) -> str:
    address = job.get("address") or {}
    if isinstance(address, dict):
        return address.get("full_location") or address.get("location") or "미확인"
    return "미확인"


def wanted_detail_text(detail: Dict[str, object]) -> Tuple[str, str, str, str]:
    job = detail.get("job") if isinstance(detail, dict) else {}
    body = job.get("detail") if isinstance(job, dict) else {}
    if not isinstance(body, dict):
        body = {}
    main_tasks = str(body.get("main_tasks") or "")
    requirements = str(body.get("requirements") or "")
    preferred = str(body.get("preferred_points") or body.get("preferred") or "")
    intro = str(body.get("intro") or body.get("benefits") or "")
    return main_tasks, requirements, preferred, intro


def classify_wanted_job(
    search_job: Dict[str, object], detail: Dict[str, object], checked_at: str, collect_date: str = "2026-07-05"
) -> ClassifiedResult:
    company = ((search_job.get("company") or {}).get("name") if isinstance(search_job.get("company"), dict) else None) or "미확인"
    title = str(search_job.get("position") or "미확인")
    link = f"https://www.wanted.co.kr/wd/{search_job.get('id')}"
    main_tasks, requirements, preferred, intro = wanted_detail_text(detail)
    all_text = " ".join([title, main_tasks, requirements, preferred, intro])
    annual_from = search_job.get("annual_from")
    due_time = search_job.get("due_time") or "상시/미확인"

    def excluded(reason: str) -> ClassifiedResult:
        return ClassifiedResult(
            included=None,
            excluded={
                "회사명": company,
                "공고명": title,
                "링크": link,
                "출처": "Wanted 상세 API",
                "제외 이유": reason,
                "확인일시": checked_at,
            },
        )

    if contains_any(title, TITLE_EXCLUDE_TERMS):
        return excluded("제목상 리드/시니어/매니저 성격이 강함")
    if contains_any(title, NON_DEVELOPMENT_TITLE_TERMS):
        return excluded("제목상 개발 직무가 아닌 운영/마케팅/기술지원 성격이 강함")
    if explicit_five_plus(all_text) or explicit_year_floor_at_least(all_text, 5) or (isinstance(annual_from, int) and annual_from >= 5):
        return excluded("상세 본문 또는 검색 메타에서 5년 이상 요구 확인")
    if contains_any(title, NON_BACKEND_CENTERED_TERMS) and not contains_any(all_text, ["backend", "백엔드", "server", "서버", "api"]):
        return excluded("프론트/모바일/연구 중심으로 백엔드 전환 가치가 불명확")
    if not contains_any(all_text, BACKEND_TERMS):
        return excluded("상세 본문에서 백엔드/API/서버/플랫폼 연관성이 부족")
    if not main_tasks and not requirements:
        return excluded("상세 본문 주요업무/자격요건 확인 불가")

    challenge = explicit_four_plus(all_text) or (isinstance(annual_from, int) and annual_from == 3)
    junior = contains_any(all_text, JUNIOR_TERMS) or (isinstance(annual_from, int) and annual_from <= 2)

    if challenge:
        fit = "중"
        priority = "중"
        decision = "도전 지원"
        memo = "요구 경력은 높지만 백엔드 업무/스택이 맞아 도전 지원 후보"
    elif junior:
        fit = "상"
        priority = "상"
        decision = "오늘 지원"
        memo = "2년차 백엔드 기준으로 경력 문턱과 업무 연관성이 좋음"
    else:
        fit = "중"
        priority = "중"
        decision = "저장"
        memo = "상세 업무는 백엔드에 맞지만 경력 적합성은 추가 확인 필요"

    backend_fit = "상" if contains_any(all_text, ["spring", "java", "kotlin", "api", "server", "서버", "백엔드"]) else "중"
    job_class = "Backend Engineer"
    if contains_any(all_text, ["devops", "infra", "인프라", "cloud", "클라우드"]):
        job_class = "Platform/DevOps"
    if contains_any(title, ["데이터 엔지니어", "data engineer"]):
        job_class = "Data Engineer"

    included = {
        "수집일": collect_date,
        "우선순위": priority,
        "회사명": company,
        "공고명": title,
        "직무분류": job_class,
        "경력조건": career_range(search_job),
        "2년차 적합도": fit,
        "백엔드 적합도": backend_fit,
        "고용형태": "미확인",
        "지역/근무형태": location_text(search_job),
        "마감일": due_time,
        "주요업무 요약": summarize_text(main_tasks),
        "필수기술": summarize_text(requirements),
        "우대기술": summarize_text(preferred),
        "지원 판단": decision,
        "지원 메모": memo,
        "출처": "Wanted",
        "링크": link,
        "확인 수준": "플랫폼 상세",
        "확인일시": checked_at,
    }
    return ClassifiedResult(included=included, excluded=None)


def wanted_candidate_score(job: Dict[str, object]) -> Tuple[int, int, int, int]:
    title = str(job.get("position") or "")
    annual_from = job.get("annual_from")
    annual_to = job.get("annual_to")
    junior_bonus = 0 if contains_any(title, JUNIOR_TERMS) else 1
    backend_bonus = 0 if contains_any(title, ["백엔드", "서버", "backend", "server", "spring", "java"]) else 1
    return (annual_from if isinstance(annual_from, int) else 9, junior_bonus, backend_bonus, annual_to if isinstance(annual_to, int) else 99)


def is_wanted_candidate(job: Dict[str, object]) -> bool:
    title = str(job.get("position") or "")
    annual_from = job.get("annual_from")
    if job.get("status") != "active" or job.get("hidden"):
        return False
    if contains_any(title, TITLE_EXCLUDE_TERMS):
        return False
    if contains_any(title, NON_DEVELOPMENT_TITLE_TERMS):
        return False
    if not contains_any(title, BACKEND_TERMS):
        return False
    if isinstance(annual_from, int) and annual_from >= 5:
        return False
    return annual_from is None or annual_from <= 3 or contains_any(title, JUNIOR_TERMS)


def fetch_json(session: requests.Session, url: str, params: Optional[Dict[str, object]] = None) -> Dict[str, object]:
    response = session.get(url, params=params, timeout=20)
    response.raise_for_status()
    return response.json()


def scan_official_sources(session: requests.Session) -> List[str]:
    summaries = []
    for name, url in OFFICIAL_SOURCES:
        try:
            response = session.get(url, timeout=15)
            state = f"{name} {response.status_code}"
            if response.ok and contains_any(response.text[:200000], BACKEND_TERMS):
                state += " 백엔드 키워드 감지"
            else:
                state += " 상세 후보 미확인"
            summaries.append(state)
        except requests.RequestException as exc:
            summaries.append(f"{name} 접근 실패: {exc.__class__.__name__}")
    return summaries


def collect_wanted_jobs(
    session: requests.Session,
    collect_date: str,
    checked_at: str,
    max_details: int,
    researched_urls: Optional[Set[str]] = None,
) -> CollectionResult:
    dedup: Dict[object, Dict[str, object]] = {}
    for keyword in KEYWORDS:
        payload = fetch_json(
            session,
            WANTED_SEARCH_URL,
            params={
                "country": "kr",
                "job_sort": "job.latest_order",
                "locations": "all",
                "years": -1,
                "limit": 50,
                "offset": 0,
                "query": keyword,
            },
        )
        for job in payload.get("data", []):
            if isinstance(job, dict) and is_wanted_candidate(job):
                dedup[job.get("id")] = job

    collection = CollectionResult([], [], 0, 0, {})
    researched_urls = researched_urls or set()
    candidates: List[Dict[str, object]] = []
    for search_job in sorted(dedup.values(), key=wanted_candidate_score):
        link = f"https://www.wanted.co.kr/wd/{search_job.get('id')}"
        if was_researched(link, researched_urls):
            collection.skipped_researched += 1
            continue
        candidates.append(search_job)
        if len(candidates) >= max_details:
            break
    for search_job in candidates:
        link = f"https://www.wanted.co.kr/wd/{search_job.get('id')}"
        try:
            detail = fetch_json(session, WANTED_DETAIL_URL.format(job_id=search_job.get("id")))
            result = classify_wanted_job(search_job, detail, checked_at, collect_date)
            collection.checked += 1
            append_classification(collection, result, link, "Wanted", checked_at)
        except requests.RequestException as exc:
            collection.checked += 1
            collection.excluded.append(
                {
                    "회사명": ((search_job.get("company") or {}).get("name") if isinstance(search_job.get("company"), dict) else "미확인"),
                    "공고명": search_job.get("position") or "미확인",
                    "링크": f"https://www.wanted.co.kr/wd/{search_job.get('id')}",
                    "출처": "Wanted 상세 API",
                    "제외 이유": f"상세 페이지 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
    return collection


def rallit_link(position_id: object) -> str:
    return f"https://www.rallit.com/positions/{position_id}"


def rallit_jobs_text(detail: Dict[str, object]) -> str:
    jobs = detail.get("jobs") or []
    if not isinstance(jobs, list):
        return ""
    return " ".join(str(job.get("name", "")) for job in jobs if isinstance(job, dict))


def is_rallit_candidate(item: Dict[str, object]) -> bool:
    title = str(item.get("title") or "")
    status = item.get("status") or {}
    if isinstance(status, dict) and status.get("code") not in {None, "HIRING"}:
        return False
    if contains_any(title, TITLE_EXCLUDE_TERMS) or contains_any(title, NON_DEVELOPMENT_TITLE_TERMS):
        return False
    text = " ".join(
        [
            title,
            " ".join(str(skill) for skill in item.get("jobSkillKeywords") or []),
            " ".join(str(level) for level in item.get("jobLevels") or []),
        ]
    )
    return contains_any(text, BACKEND_TERMS)


def classify_rallit_position(
    list_item: Dict[str, object], detail: Dict[str, object], checked_at: str, collect_date: str = "2026-07-05"
) -> ClassifiedResult:
    title = str(detail.get("title") or list_item.get("title") or "미확인")
    company = str(detail.get("companyName") or list_item.get("companyName") or "미확인")
    position_id = detail.get("id") or list_item.get("id")
    link = str(list_item.get("url") or rallit_link(position_id))
    ended_at = detail.get("endedAt") or list_item.get("endedAt") or "상시/미확인"
    levels = detail.get("jobLevels") or list_item.get("jobLevels") or []
    skill_keywords = detail.get("jobSkillKeywords") or list_item.get("jobSkillKeywords") or []

    responsibilities = html_to_text(detail.get("responsibilities"))
    requirements = html_to_text(detail.get("basicQualifications"))
    preferred = html_to_text(detail.get("preferredQualifications"))
    job_names = rallit_jobs_text(detail)
    all_text = " ".join([title, job_names, responsibilities, requirements, preferred, " ".join(map(str, skill_keywords))])

    def excluded(reason: str) -> ClassifiedResult:
        return ClassifiedResult(
            included=None,
            excluded={
                "회사명": company,
                "공고명": title,
                "링크": link,
                "출처": "Rallit 상세 API",
                "제외 이유": reason,
                "확인일시": checked_at,
            },
        )

    status = detail.get("status") or list_item.get("status") or {}
    if isinstance(status, dict) and status.get("code") not in {None, "HIRING"}:
        return excluded("모집 중 상태가 아님")
    if contains_any(title, TITLE_EXCLUDE_TERMS) or contains_any(title, NON_DEVELOPMENT_TITLE_TERMS):
        return excluded("제목상 리드/비개발 직무 성격이 강함")
    if title_has_non_backend_focus(title):
        return excluded("프론트/모바일 중심으로 백엔드 전환 가치가 불명확")
    if explicit_five_plus(all_text) or explicit_year_floor_at_least(all_text, 5):
        return excluded("상세 본문에서 5년 이상 요구 확인")
    if not contains_any(all_text, BACKEND_TERMS):
        return excluded("상세 본문에서 백엔드/API/서버/플랫폼 연관성이 부족")
    if not responsibilities and not requirements:
        return excluded("상세 본문 주요업무/자격요건 확인 불가")

    level_text = " ".join(str(level) for level in levels)
    beginner = contains_any(level_text, ["BEGINNER", "IRRELEVANT"]) or contains_any(all_text, JUNIOR_TERMS)
    challenge = explicit_four_plus(all_text) or re.search(r"3\s*년\s*이상", all_text)
    if challenge and not beginner:
        priority = "중"
        fit = "중"
        decision = "도전 지원"
        memo = "요구 경력은 높지만 백엔드 업무/스택이 맞아 도전 지원 후보"
    else:
        priority = "상"
        fit = "상"
        decision = "오늘 지원"
        memo = "랠릿 상세에서 신입/주니어 또는 2년차 접근 가능한 백엔드 업무 확인"

    backend_fit = "상" if contains_any(all_text, ["spring", "java", "kotlin", "django", "api", "server", "서버", "백엔드"]) else "중"
    job_class = "Backend Engineer"
    if contains_any(all_text, ["devops", "infra", "인프라", "cloud", "클라우드"]):
        job_class = "Platform/DevOps"
    if contains_any(all_text, ["데이터 엔지니어", "data engineer"]):
        job_class = "Data Engineer"

    address_parts = [detail.get("addressMain"), detail.get("addressDetail")]
    included = {
        "수집일": collect_date,
        "우선순위": priority,
        "회사명": company,
        "공고명": title,
        "직무분류": job_class,
        "경력조건": "신입/주니어" if beginner else "미확인",
        "2년차 적합도": fit,
        "백엔드 적합도": backend_fit,
        "고용형태": "미확인",
        "지역/근무형태": clean_text(" ".join(str(part) for part in address_parts if part)) or "미확인",
        "마감일": ended_at,
        "주요업무 요약": summarize_text(responsibilities),
        "필수기술": summarize_text(requirements),
        "우대기술": summarize_text(preferred or ", ".join(map(str, skill_keywords))),
        "지원 판단": decision,
        "지원 메모": memo,
        "출처": "Rallit",
        "링크": link,
        "확인 수준": "플랫폼 상세",
        "확인일시": checked_at,
    }
    return ClassifiedResult(included=included, excluded=None)


def collect_rallit_jobs(
    session: requests.Session,
    collect_date: str,
    checked_at: str,
    max_details: int,
    researched_urls: Optional[Set[str]] = None,
) -> CollectionResult:
    collection = CollectionResult([], [], 0, 0, {})
    researched_urls = researched_urls or set()
    candidates: List[Dict[str, object]] = []
    page_number = 1
    while len(candidates) < max_details:
        payload = fetch_json(session, RALLIT_LIST_URL, params={"pageNumber": page_number, "pageSize": 60})
        items = (((payload.get("data") or {}).get("items")) if isinstance(payload.get("data"), dict) else []) or []
        if not items:
            break
        for item in items:
            if not isinstance(item, dict) or not is_rallit_candidate(item):
                continue
            link = str(item.get("url") or rallit_link(item.get("id")))
            if was_researched(link, researched_urls):
                collection.skipped_researched += 1
                continue
            candidates.append(item)
            if len(candidates) >= max_details:
                break
        page_number += 1
    for item in candidates:
        link = str(item.get("url") or rallit_link(item.get("id")))
        try:
            detail_payload = fetch_json(session, RALLIT_DETAIL_URL.format(position_id=item.get("id")))
            detail = detail_payload.get("data") if isinstance(detail_payload.get("data"), dict) else {}
            result = classify_rallit_position(item, detail, checked_at, collect_date)
            collection.checked += 1
            append_classification(collection, result, link, "Rallit", checked_at)
        except requests.RequestException as exc:
            collection.checked += 1
            collection.excluded.append(
                {
                    "회사명": item.get("companyName") or "미확인",
                    "공고명": item.get("title") or "미확인",
                    "링크": item.get("url") or rallit_link(item.get("id")),
                    "출처": "Rallit 상세 API",
                    "제외 이유": f"상세 페이지 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
    return collection


def jumpit_link(position_id: object) -> str:
    return f"https://jumpit.saramin.co.kr/position/{position_id}"


def is_jumpit_candidate(item: Dict[str, object]) -> bool:
    title = str(item.get("title") or "")
    if contains_any(title, TITLE_EXCLUDE_TERMS) or contains_any(title, NON_DEVELOPMENT_TITLE_TERMS):
        return False
    tech_stacks = item.get("techStacks") or []
    text = " ".join(
        [
            title,
            str(item.get("jobCategory") or ""),
            " ".join(str(stack.get("stack") if isinstance(stack, dict) else stack) for stack in tech_stacks),
        ]
    )
    return contains_any(text, BACKEND_TERMS)


def classify_jumpit_position(
    list_item: Dict[str, object], detail: Dict[str, object], checked_at: str, collect_date: str = "2026-07-05"
) -> ClassifiedResult:
    title = str(detail.get("title") or list_item.get("title") or "미확인")
    company = str(detail.get("companyName") or list_item.get("companyName") or "미확인")
    position_id = detail.get("id") or list_item.get("id")
    link = jumpit_link(position_id)
    tech_stacks = detail.get("techStacks") or list_item.get("techStacks") or []
    tech_text = ", ".join(str(stack.get("stack") if isinstance(stack, dict) else stack) for stack in tech_stacks)
    responsibilities = html_to_text(detail.get("responsibility"))
    requirements = html_to_text(detail.get("qualifications"))
    preferred = html_to_text(detail.get("preferredRequirements"))
    all_text = " ".join([title, responsibilities, requirements, preferred, tech_text])
    min_career = detail.get("minCareer", list_item.get("minCareer"))
    max_career = detail.get("maxCareer", list_item.get("maxCareer"))

    def excluded(reason: str) -> ClassifiedResult:
        return ClassifiedResult(
            included=None,
            excluded={
                "회사명": company,
                "공고명": title,
                "링크": link,
                "출처": "Jumpit 상세 API",
                "제외 이유": reason,
                "확인일시": checked_at,
            },
        )

    if contains_any(title, TITLE_EXCLUDE_TERMS) or contains_any(title, NON_DEVELOPMENT_TITLE_TERMS):
        return excluded("제목상 리드/비개발 직무 성격이 강함")
    if explicit_five_plus(all_text) or explicit_year_floor_at_least(all_text, 5) or (
        isinstance(min_career, int) and min_career >= 5
    ):
        return excluded("상세 본문 또는 경력 조건에서 5년 이상 요구 확인")
    if title_has_non_backend_focus(title):
        return excluded("프론트/모바일 중심으로 백엔드 전환 가치가 불명확")
    if not contains_any(all_text, BACKEND_TERMS):
        return excluded("상세 본문에서 백엔드/API/서버/플랫폼 연관성이 부족")
    if not responsibilities and not requirements:
        return excluded("상세 본문 주요업무/자격요건 확인 불가")

    career_text = career_range({"annual_from": min_career, "annual_to": max_career})
    challenge = explicit_four_plus(all_text) or (isinstance(min_career, int) and min_career >= 3)
    junior = contains_any(all_text, JUNIOR_TERMS) or (isinstance(min_career, int) and min_career <= 2)
    if challenge and not junior:
        priority, fit, decision = "중", "중", "도전 지원"
        memo = "점핏 상세에서 3년 이상 경력 성격이 보여 도전 지원 후보"
    elif junior:
        priority, fit, decision = "상", "상", "오늘 지원"
        memo = "점핏 상세에서 2년차에 맞는 백엔드 경력 조건과 업무를 확인"
    else:
        priority, fit, decision = "중", "중", "저장"
        memo = "점핏 상세 업무는 백엔드에 맞지만 경력 조건은 추가 확인 필요"

    locations = detail.get("locations") or list_item.get("locations") or []
    location = clean_text(", ".join(str(value) for value in locations if value)) if isinstance(locations, list) else clean_text(locations)
    deadline = str(detail.get("closedAt") or list_item.get("closedAt") or "상시/미확인")
    included = {
        "수집일": collect_date,
        "우선순위": priority,
        "회사명": company,
        "공고명": title,
        "직무분류": "Backend Engineer",
        "경력조건": career_text,
        "2년차 적합도": fit,
        "백엔드 적합도": "상" if contains_any(all_text, ["spring", "java", "kotlin", "api", "server", "서버", "백엔드"]) else "중",
        "고용형태": clean_text(detail.get("employmentType")) or "미확인",
        "지역/근무형태": location or "미확인",
        "마감일": deadline[:10] if re.match(r"\d{4}-\d{2}-\d{2}", deadline) else deadline,
        "주요업무 요약": summarize_text(responsibilities),
        "필수기술": summarize_text(requirements),
        "우대기술": summarize_text(preferred or tech_text),
        "지원 판단": decision,
        "지원 메모": memo,
        "출처": "Jumpit",
        "링크": link,
        "확인 수준": "플랫폼 상세",
        "확인일시": checked_at,
    }
    return ClassifiedResult(included=included, excluded=None)


def collect_jumpit_jobs(
    session: requests.Session,
    collect_date: str,
    checked_at: str,
    max_details: int,
    researched_urls: Optional[Set[str]] = None,
) -> CollectionResult:
    collection = CollectionResult([], [], 0, 0, {})
    researched_urls = researched_urls or set()
    candidates: List[Dict[str, object]] = []
    seen_ids = set()
    page = 1
    while len(candidates) < max_details:
        payload = fetch_json(session, JUMPIT_LIST_URL, params={"sort": "latest", "page": page})
        result = payload.get("result") if isinstance(payload.get("result"), dict) else {}
        positions = result.get("positions") if isinstance(result, dict) else []
        if not isinstance(positions, list) or not positions:
            break
        new_ids = 0
        for item in positions:
            if not isinstance(item, dict) or item.get("id") in seen_ids:
                continue
            seen_ids.add(item.get("id"))
            new_ids += 1
            if not is_jumpit_candidate(item):
                continue
            link = jumpit_link(item.get("id"))
            if was_researched(link, researched_urls):
                collection.skipped_researched += 1
                continue
            candidates.append(item)
            if len(candidates) >= max_details:
                break
        if new_ids == 0:
            break
        page += 1

    for item in candidates:
        link = jumpit_link(item.get("id"))
        try:
            detail_payload = fetch_json(session, JUMPIT_DETAIL_URL.format(position_id=item.get("id")))
            detail = detail_payload.get("result") if isinstance(detail_payload.get("result"), dict) else {}
            result = classify_jumpit_position(item, detail, checked_at, collect_date)
            collection.checked += 1
            append_classification(collection, result, link, "Jumpit", checked_at)
        except requests.RequestException as exc:
            collection.checked += 1
            collection.excluded.append(
                {
                    "회사명": item.get("companyName") or "미확인",
                    "공고명": item.get("title") or "미확인",
                    "링크": link,
                    "출처": "Jumpit 상세 API",
                    "제외 이유": f"상세 페이지 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
    return collection


def enrich_html_classification(
    result: ClassifiedResult,
    *,
    company: str = "",
    title: str = "",
    career: str = "",
    location: str = "",
    deadline: str = "",
    employment_type: str = "",
    summary: str = "",
) -> ClassifiedResult:
    fields = {
        "회사명": clean_text(company),
        "공고명": clean_text(title),
        "경력조건": clean_text(career),
        "지역/근무형태": clean_text(location),
        "마감일": clean_text(deadline),
        "고용형태": clean_text(employment_type),
        "주요업무 요약": summarize_text(summary) if summary else "",
    }
    target = result.included if result.included is not None else result.excluded
    if target is None:
        return result
    for key, value in fields.items():
        if value:
            target[key] = value
    return result


def listing_has_backend_signal(item: Dict[str, object], fields: Sequence[str]) -> bool:
    text = " ".join(str(item.get(field) or "") for field in fields)
    if contains_any(text, TITLE_EXCLUDE_TERMS) or contains_any(text, NON_DEVELOPMENT_TITLE_TERMS):
        return False
    return contains_any(
        text,
        [
            "백엔드",
            "서버",
            "backend",
            "server",
            "spring",
            "java",
            "kotlin",
            "api",
            "devops",
            "플랫폼 엔지니어",
            "플랫폼 개발",
            "platform engineer",
            "platform developer",
            "cloud engineer",
            "클라우드 엔지니어",
            "infra engineer",
            "인프라 엔지니어",
            "sre",
            "site reliability",
            "데이터 엔지니어",
            "data engineer",
            "데이터베이스",
            "database",
            "dba",
        ],
    )


def catch_link(recruit_id: object) -> str:
    return f"https://www.catch.co.kr/NCS/RecruitInfoDetails/{recruit_id}"


def catch_summary(html_text: str) -> str:
    match = re.search(r'"description":"([^"\\]*(?:\\.[^"\\]*)*)"', html_text)
    if not match:
        return ""
    try:
        return json.loads(f'"{match.group(1)}"')
    except json.JSONDecodeError:
        return ""


def collect_catch_jobs(
    session: requests.Session,
    collect_date: str,
    checked_at: str,
    max_details: int,
    researched_urls: Optional[Set[str]] = None,
) -> CollectionResult:
    collection = CollectionResult([], [], 0, 0, {})
    researched_urls = researched_urls or set()
    candidates: List[Dict[str, object]] = []
    seen_ids = set()
    page = 1
    while len(candidates) < max_details:
        payload = fetch_json(
            session,
            CATCH_LIST_URL,
            params={"pageSize": 30, "curpage": page, "Sort": 0, "onRecruitYN": "Y"},
        )
        items = payload.get("recruitData") if isinstance(payload.get("recruitData"), list) else []
        if not items:
            break
        new_ids = 0
        for item in items:
            if not isinstance(item, dict) or item.get("RecruitID") in seen_ids:
                continue
            seen_ids.add(item.get("RecruitID"))
            new_ids += 1
            if not listing_has_backend_signal(item, ["RecruitTitle", "Depth", "AssignedTaskNameListString"]):
                continue
            link = catch_link(item.get("RecruitID"))
            if was_researched(link, researched_urls):
                collection.skipped_researched += 1
                continue
            candidates.append(item)
            if len(candidates) >= max_details:
                break
        if new_ids == 0:
            break
        page += 1

    for item in candidates:
        link = catch_link(item.get("RecruitID"))
        try:
            response = session.get(link, timeout=20)
            response.raise_for_status()
            result = classify_html_platform_detail("Catch", "캐치", link, response.text, checked_at, collect_date)
            career = " ".join(
                part for part in [str(item.get("CareerGubunCode") or ""), str(item.get("ExperienceRange") or "")] if part
            )
            deadline = str(item.get("ApplyEndCode") or item.get("ApplyEndDatetime") or "상시/미확인")
            result = enrich_html_classification(
                result,
                company=str(item.get("CompName") or ""),
                title=str(item.get("RecruitTitle") or ""),
                career=career,
                location=str(item.get("WorkArea") or ""),
                deadline=deadline[:10] if re.match(r"\d{4}-\d{2}-\d{2}", deadline) else deadline,
                employment_type=str(item.get("GubunCode") or ""),
                summary=catch_summary(response.text),
            )
            collection.checked += 1
            append_classification(collection, result, link, "Catch", checked_at)
        except requests.RequestException as exc:
            collection.checked += 1
            collection.excluded.append(
                {
                    "회사명": item.get("CompName") or "미확인",
                    "공고명": item.get("RecruitTitle") or "미확인",
                    "링크": link,
                    "출처": "Catch 상세 HTML",
                    "제외 이유": f"상세 페이지 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
    return collection


def incruit_html(response: requests.Response) -> str:
    try:
        return response.content.decode("euc-kr")
    except UnicodeDecodeError:
        return response.text


def extract_incruit_company_title(html_text: str) -> Tuple[str, str]:
    soup = BeautifulSoup(html_text, "lxml")
    meta_title = meta_content(soup, 'meta[name="title"]')
    if not meta_title:
        return "", ""
    title = re.sub(r"\s*[:|]\s*인크루트.*$", "", meta_title)
    if "," in title:
        company, job = title.split(",", 1)
        return clean_text(company), clean_text(job)
    return "", clean_text(title)


def collect_incruit_jobs(
    session: requests.Session,
    collect_date: str,
    checked_at: str,
    max_details: int,
    researched_urls: Optional[Set[str]] = None,
) -> CollectionResult:
    search_results: List[List[str]] = []
    for keyword in platform_search_terms():
        response = session.get(INCRUIT_SEARCH_URL, params={"col": "job", "kw": keyword}, timeout=20)
        response.raise_for_status()
        links = []
        for href in re.findall(r'href=["\']([^"\']*jobdb_info/jobpost\.asp\?job=[^"\']+)', response.text, re.IGNORECASE):
            url = absolute_url("https://job.incruit.com", href)
            if url not in links:
                links.append(url)
        search_results.append(links)
    collection = CollectionResult([], [], 0, 0, {})
    researched_urls = researched_urls or set()
    candidates = []
    for link in merge_platform_links(search_results, max_details * 3):
        if was_researched(link, researched_urls):
            collection.skipped_researched += 1
            continue
        candidates.append(link)
        if len(candidates) >= max_details:
            break
    for link in candidates:
        try:
            response = session.get(link, timeout=20)
            response.raise_for_status()
            detail_html = incruit_html(response)
            result = classify_html_platform_detail("Incruit", "인크루트", link, detail_html, checked_at, collect_date)
            company, title = extract_incruit_company_title(detail_html)
            result = enrich_html_classification(result, company=company, title=title)
            collection.checked += 1
            append_classification(collection, result, link, "Incruit", checked_at)
        except requests.RequestException as exc:
            collection.checked += 1
            collection.excluded.append(
                {
                    "회사명": "미확인",
                    "공고명": "인크루트 상세 접근 실패",
                    "링크": link,
                    "출처": "Incruit 상세 HTML",
                    "제외 이유": f"상세 페이지 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
    return collection


def worknet_link(wanted_auth_no: str) -> str:
    return (
        "https://www.work24.go.kr/wk/a/b/1500/empDetailAuthView.do?"
        f"wantedAuthNo={wanted_auth_no}&infoTypeCd=VALIDATION&infoTypeGroup=tb_workinfoworknet"
    )


def worknet_detail_fields(html_text: str) -> Tuple[str, str, str, str, str, str, str]:
    soup = BeautifulSoup(html_text, "lxml")
    company = clean_text((soup.select_one("p.corp_info strong") or {}).get_text(" ", strip=True) if soup.select_one("p.corp_info strong") else "")
    title = clean_text((soup.select_one("strong.title") or {}).get_text(" ", strip=True) if soup.select_one("strong.title") else "")
    table_fields: Dict[str, str] = {}
    for row in soup.select("tr"):
        header = clean_text((row.select_one("th") or {}).get_text(" ", strip=True) if row.select_one("th") else "")
        value = clean_text((row.select_one("td") or {}).get_text(" ", strip=True) if row.select_one("td") else "")
        if header and value:
            table_fields[header] = value
    for item in soup.select("li"):
        header = clean_text((item.select_one("em.tit") or {}).get_text(" ", strip=True) if item.select_one("em.tit") else "")
        value = clean_text((item.select_one("p") or {}).get_text(" ", strip=True) if item.select_one("p") else "")
        if header and value:
            table_fields[header] = value
    deadline = ""
    for heading in soup.find_all(["strong", "em"]):
        if clean_text(heading.get_text(" ", strip=True)) != "접수 마감일":
            continue
        sibling = heading.find_next("p")
        deadline = clean_text(sibling.get_text(" ", strip=True) if sibling else "")
        if deadline:
            break
    task = clean_text((soup.select_one("div.fold") or {}).get_text(" ", strip=True) if soup.select_one("div.fold") else "")
    return (
        company,
        title,
        table_fields.get("경력", ""),
        table_fields.get("근무지역", table_fields.get("근무예정지", table_fields.get("지역", ""))),
        deadline or table_fields.get("접수마감일", ""),
        table_fields.get("고용형태", ""),
        task,
    )


def collect_worknet_jobs(
    session: requests.Session,
    collect_date: str,
    checked_at: str,
    max_details: int,
    researched_urls: Optional[Set[str]] = None,
) -> CollectionResult:
    search_results: List[List[str]] = []
    for keyword in platform_search_terms():
        response = session.get(
            WORKNET_SEARCH_URL,
            params={"srcKeyword": keyword, "keywordWantedTitle": "keywordWantedTitle"},
            timeout=20,
        )
        response.raise_for_status()
        links = []
        for wanted_auth_no in re.findall(r"wantedAuthNo=([^&\"']+)", html.unescape(response.text)):
            url = worknet_link(wanted_auth_no)
            if url not in links:
                links.append(url)
        search_results.append(links)
    collection = CollectionResult([], [], 0, 0, {})
    researched_urls = researched_urls or set()
    candidates = []
    for link in merge_platform_links(search_results, max_details * 3):
        if was_researched(link, researched_urls):
            collection.skipped_researched += 1
            continue
        candidates.append(link)
        if len(candidates) >= max_details:
            break
    for link in candidates:
        try:
            response = session.get(link, timeout=20)
            response.raise_for_status()
            result = classify_html_platform_detail("Worknet", "고용24", link, response.text, checked_at, collect_date)
            company, title, career, location, deadline, employment_type, task = worknet_detail_fields(response.text)
            result = enrich_html_classification(
                result,
                company=company,
                title=title,
                career=career,
                location=location,
                deadline=deadline,
                employment_type=employment_type,
                summary=task,
            )
            collection.checked += 1
            append_classification(collection, result, link, "Worknet", checked_at)
        except requests.RequestException as exc:
            collection.checked += 1
            collection.excluded.append(
                {
                    "회사명": "미확인",
                    "공고명": "고용24 상세 접근 실패",
                    "링크": link,
                    "출처": "Worknet 상세 HTML",
                    "제외 이유": f"상세 페이지 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
    return collection


def starting_link(public_number: object) -> str:
    return STARTING_DETAIL_URL.format(public_number=public_number)


def collect_starting_jobs(
    session: requests.Session,
    collect_date: str,
    checked_at: str,
    max_details: int,
    researched_urls: Optional[Set[str]] = None,
) -> CollectionResult:
    collection = CollectionResult([], [], 0, 0, {})
    researched_urls = researched_urls or set()
    candidates: List[Dict[str, object]] = []
    offset = 0
    while len(candidates) < max_details:
        payload = fetch_json(session, STARTING_LIST_URL, params={"offset": offset, "limit": 100})
        items = payload.get("rows") if isinstance(payload.get("rows"), list) else []
        if not items:
            break
        for item in items:
            if not isinstance(item, dict) or not listing_has_backend_signal(item, ["title"]):
                continue
            link = starting_link(item.get("publicNumber"))
            if was_researched(link, researched_urls):
                collection.skipped_researched += 1
                continue
            candidates.append(item)
            if len(candidates) >= max_details:
                break
        if len(items) < 100:
            break
        offset += len(items)
    for item in candidates:
        link = starting_link(item.get("publicNumber"))
        try:
            response = session.get(link, timeout=20)
            response.raise_for_status()
            result = classify_html_platform_detail("Starting", "스타팅", link, response.text, checked_at, collect_date)
            result = enrich_html_classification(
                result,
                company=str(item.get("companyName") or ""),
                title=str(item.get("title") or ""),
                career=str(item.get("careerLabel") or ""),
                location=str(item.get("workAddressShort") or ""),
                employment_type=str(item.get("contractType") or item.get("employmentType") or ""),
            )
            collection.checked += 1
            append_classification(collection, result, link, "Starting", checked_at)
        except requests.RequestException as exc:
            collection.checked += 1
            collection.excluded.append(
                {
                    "회사명": item.get("companyName") or "미확인",
                    "공고명": item.get("title") or "미확인",
                    "링크": link,
                    "출처": "Starting 상세 HTML",
                    "제외 이유": f"상세 페이지 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
    return collection


def absolute_url(base: str, href: str) -> str:
    href = html.unescape(href)
    if href.startswith("http"):
        return href
    return base.rstrip("/") + "/" + href.lstrip("/")


def collect_saramin_jobs(
    session: requests.Session,
    collect_date: str,
    checked_at: str,
    max_details: int,
    researched_urls: Optional[Set[str]] = None,
) -> CollectionResult:
    search_results: List[List[str]] = []
    for keyword in platform_search_terms():
        term_links: List[str] = []
        response = session.get(SARAMIN_SEARCH_URL, params={"searchword": keyword}, timeout=20)
        response.raise_for_status()
        for href in re.findall(r'href=["\']([^"\']*zf_user/jobs/relay/view[^"\']+)["\']', response.text):
            url = absolute_url("https://www.saramin.co.kr", href)
            if url not in term_links:
                term_links.append(url)
        search_results.append(term_links)
    collection = CollectionResult([], [], 0, 0, {})
    researched_urls = researched_urls or set()
    links = merge_platform_links(search_results, max_details * 3)
    candidates: List[str] = []
    for url in links:
        if was_researched(url, researched_urls):
            collection.skipped_researched += 1
            continue
        candidates.append(url)
        if len(candidates) >= max_details:
            break
    for url in candidates:
        try:
            detail_response = session.get(url, timeout=20)
            detail_response.raise_for_status()
            result = classify_saramin_detail(url, detail_response.text, checked_at, collect_date)
            collection.checked += 1
            append_classification(collection, result, url, "Saramin", checked_at)
        except requests.RequestException as exc:
            collection.checked += 1
            collection.excluded.append(
                {
                    "회사명": "미확인",
                    "공고명": "사람인 상세 접근 실패",
                    "링크": url,
                    "출처": "Saramin 상세 HTML",
                    "제외 이유": f"상세 페이지 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
    return collection


def collect_jobkorea_jobs(
    session: requests.Session,
    collect_date: str,
    checked_at: str,
    max_details: int,
    researched_urls: Optional[Set[str]] = None,
) -> CollectionResult:
    search_results: List[List[str]] = []
    for keyword in platform_search_terms():
        term_links: List[str] = []
        response = session.get(JOBKOREA_SEARCH_URL, params={"stext": keyword}, timeout=20)
        response.raise_for_status()
        for href in re.findall(r'https://www\.jobkorea\.co\.kr/Recruit/GI_Read/\d+[^"\']*', response.text):
            url = html.unescape(href)
            if url not in term_links:
                term_links.append(url)
        search_results.append(term_links)
    collection = CollectionResult([], [], 0, 0, {})
    researched_urls = researched_urls or set()
    links = merge_platform_links(search_results, max_details * 3)
    candidates: List[str] = []
    for url in links:
        if was_researched(url, researched_urls):
            collection.skipped_researched += 1
            continue
        candidates.append(url)
        if len(candidates) >= max_details:
            break
    for url in candidates:
        try:
            detail_response = session.get(url, timeout=20)
            detail_response.raise_for_status()
            result = classify_jobkorea_detail(url, detail_response.text, checked_at, collect_date)
            collection.checked += 1
            append_classification(collection, result, url, "JobKorea", checked_at)
        except requests.RequestException as exc:
            collection.checked += 1
            collection.excluded.append(
                {
                    "회사명": "미확인",
                    "공고명": "잡코리아 상세 접근 실패",
                    "링크": url,
                    "출처": "JobKorea 상세 HTML",
                    "제외 이유": f"상세 페이지 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
    return collection


def days_until_deadline(value: object, collect_date: str) -> Optional[int]:
    if not value or value in {"상시/미확인", "미확인"}:
        return None
    text = clean_text(value)
    if "오늘" in text or re.search(r"(?:^|[^a-z0-9])d\s*-\s*0(?:$|[^a-z0-9])", text.lower()):
        return 0
    if "내일" in text or re.search(r"(?:^|[^a-z0-9])d\s*-\s*1(?:$|[^a-z0-9])", text.lower()):
        return 1
    try:
        normalized = re.sub(r"[./]", "-", text[:10])
        return (date.fromisoformat(normalized) - date.fromisoformat(collect_date)).days
    except ValueError:
        return None


def sort_jobs(jobs: Sequence[Dict[str, object]], collect_date: str) -> List[Dict[str, object]]:
    priority_order = {"상": 0, "중": 1, "하": 2}
    fit_order = {"상": 0, "중": 1, "하": 2}

    def key(job: Dict[str, object]) -> Tuple[int, int, int, str]:
        days = days_until_deadline(job.get("마감일"), collect_date)
        return (
            priority_order.get(str(job.get("우선순위")), 9),
            fit_order.get(str(job.get("2년차 적합도")), 9),
            days if days is not None else 999,
            str(job.get("회사명")),
        )

    return sorted(jobs, key=key)


def job_identity(job: Dict[str, object]) -> Tuple[str, str, str]:
    company = clean_text(job.get("회사명")).lower()
    title = clean_text(job.get("공고명")).lower()
    if company and company != "미확인" and title and title != "미확인":
        return ("company_title", company, title)
    link = clean_text(job.get("링크")).split("?", 1)[0].lower()
    if not link:
        return ("row", str(id(job)), "")
    return ("link", link, "")


def previous_day_workbook_path(collect_date: str) -> Path:
    previous_date = date.fromisoformat(collect_date) - timedelta(days=1)
    return Path(f"/tmp/it-backend-jobs-{previous_date.isoformat()}.xlsx")


def read_previous_job_identities(workbook_path: Path) -> Tuple[Optional[set], str]:
    if not workbook_path.exists():
        return None, "전일 최종 XLSX 없음: 중복 비교 미실시"
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        if "채용공고" not in workbook.sheetnames:
            return None, "전일 최종 XLSX에 채용공고 시트 없음: 중복 비교 미실시"
        sheet = workbook["채용공고"]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header_index = {clean_text(value): index for index, value in enumerate(header_row) if clean_text(value)}
        required = {"회사명", "공고명", "링크"}
        if not required.issubset(header_index):
            return None, "전일 최종 XLSX 헤더 확인 실패: 중복 비교 미실시"

        identities = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            item = {
                "회사명": row[header_index["회사명"]] if len(row) > header_index["회사명"] else "",
                "공고명": row[header_index["공고명"]] if len(row) > header_index["공고명"] else "",
                "링크": row[header_index["링크"]] if len(row) > header_index["링크"] else "",
            }
            if any(clean_text(value) for value in item.values()):
                identities.add(job_identity(item))
        return identities, f"전일 최종 XLSX 중복 비교 준비: {len(identities)}건"
    except Exception as exc:  # pragma: no cover - openpyxl exception types vary by corrupt file.
        return None, f"전일 최종 XLSX 열기 실패({exc.__class__.__name__}): 중복 비교 미실시"


def filter_previous_day_duplicates(
    jobs: Sequence[Dict[str, object]],
    previous_workbook_path: Path,
    collect_date: str,
    checked_at: str,
) -> Tuple[List[Dict[str, object]], List[Dict[str, object]], str]:
    previous_identities, source_note = read_previous_job_identities(previous_workbook_path)
    if previous_identities is None:
        return list(jobs), [], source_note

    eligible: List[Dict[str, object]] = []
    excluded: List[Dict[str, object]] = []
    urgent_reincluded = 0
    for job in jobs:
        if job_identity(job) not in previous_identities:
            eligible.append(job)
            continue
        if days_until_deadline(job.get("마감일"), collect_date) in {0, 1}:
            eligible.append(job)
            urgent_reincluded += 1
            continue
        excluded.append(
            {
                "회사명": job.get("회사명", "미확인"),
                "공고명": job.get("공고명", "미확인"),
                "링크": job.get("링크", ""),
                "출처": job.get("출처", "미확인"),
                "제외 이유": "전일 포함 공고: 마감일이 오늘·내일이 아니어서 제외",
                "확인일시": checked_at,
            }
        )
    return (
        eligible,
        excluded,
        f"{source_note}; 전일 중복 제외 {len(excluded)}건 / 마감 임박 재포함 {urgent_reincluded}건",
    )


def select_jobs_for_workbook(jobs: Sequence[Dict[str, object]], collect_date: str, limit: int = 200) -> List[Dict[str, object]]:
    grouped: Dict[str, List[Dict[str, object]]] = {}
    seen_identities = set()
    for job in sort_jobs(jobs, collect_date):
        identity = job_identity(job)
        if identity in seen_identities:
            continue
        seen_identities.add(identity)
        grouped.setdefault(str(job.get("출처") or "미확인"), []).append(job)
    selected: List[Dict[str, object]] = []
    while len(selected) < limit and any(grouped.values()):
        for source in sorted(grouped):
            if grouped[source] and len(selected) < limit:
                selected.append(grouped[source].pop(0))
    return selected


def selected_job_count(jobs: Sequence[Dict[str, object]], collect_date: str) -> int:
    return len(select_jobs_for_workbook(jobs, collect_date, limit=TARGET_INCLUDED_JOBS))


def build_workbook(
    output_path: Path,
    collect_date: str,
    jobs: Sequence[Dict[str, object]],
    excluded: Sequence[Dict[str, object]],
    source_summary: Sequence[str],
    limitation_note: str,
    checked_count: Optional[int] = None,
) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "요약"
    jobs_sheet = wb.create_sheet("채용공고")
    excluded_sheet = wb.create_sheet("제외_검토")

    high_fit = sum(1 for job in jobs if job.get("2년차 적합도") == "상")
    urgent = sum(
        1
        for job in jobs
        if (days := days_until_deadline(job.get("마감일"), collect_date)) is not None and 0 <= days <= 7
    )
    checked_count = checked_count if checked_count is not None else len(jobs) + len(excluded)

    summary_rows = [
        ["항목", "값"],
        ["수집일", collect_date],
        ["전체 확인 공고 수", checked_count],
        ["엑셀에 포함한 공고 수", len(jobs)],
        ["2년차 백엔드 적합 상 공고 수", high_fit],
        ["마감 7일 이내 공고 수", urgent],
        ["오늘의 검색/확인 출처 요약", " / ".join(source_summary) or "미확인"],
        ["정확도/한계 메모", limitation_note or "상세 본문 확인 공고만 포함"],
    ]

    for row in summary_rows:
        summary.append(row)

    jobs_sheet.append(JOB_HEADERS)
    for job in jobs:
        jobs_sheet.append([job.get(JOB_SHEET_FIELDS[header], "") for header in JOB_HEADERS])

    excluded_sheet.append(EXCLUDED_HEADERS)
    for item in excluded:
        excluded_sheet.append([item.get(header, "") for header in EXCLUDED_HEADERS])

    for ws in [summary, jobs_sheet, excluded_sheet]:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 26
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, jobs_sheet.max_row + 1):
        jobs_sheet.row_dimensions[row].height = 46
        fill = PatternFill("solid", fgColor="F3F8FC" if row % 2 == 0 else "EAF2F8")
        for cell in jobs_sheet[row]:
            cell.fill = fill

    for row in range(2, excluded_sheet.max_row + 1):
        for cell in excluded_sheet[row]:
            cell.fill = PatternFill("solid", fgColor="FFF7ED")

    for row in range(2, 9):
        summary.cell(row=row, column=1).font = Font(bold=True, color="1F2937")
        summary.cell(row=row, column=1).fill = PatternFill("solid", fgColor="DCEAF7")
    widths = {
        "요약": [28, 80, 42, 70],
        "채용공고": [48, 20, 38, 16, 24, 15, 64, 16],
        "제외_검토": [18, 42, 58, 20, 82, 20],
    }
    for ws in [summary, jobs_sheet, excluded_sheet]:
        for index, width in enumerate(widths[ws.title], 1):
            ws.column_dimensions[get_column_letter(index)].width = width

    link_columns = {"요약": 4, "채용공고": 1, "제외_검토": 3}
    for ws in [summary, jobs_sheet, excluded_sheet]:
        link_col = link_columns[ws.title]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_col)
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def validate_workbook_for_delivery(output_path: Path, jobs: Sequence[Dict[str, object]]) -> List[str]:
    errors: List[str] = []
    if not output_path.exists() or output_path.stat().st_size == 0:
        errors.append("XLSX 파일이 없거나 비어 있음")
        return errors
    try:
        workbook = load_workbook(output_path, read_only=False)
    except Exception as exc:  # pragma: no cover - exact openpyxl exceptions vary.
        return [f"XLSX 열기 실패: {exc.__class__.__name__}"]

    if workbook.sheetnames != ["요약", "채용공고", "제외_검토"]:
        errors.append(f"필수 시트 구성 불일치: {workbook.sheetnames}")
    for sheet_name in ["요약", "채용공고", "제외_검토"]:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if not sheet.freeze_panes:
                errors.append(f"{sheet_name} freeze pane 누락")
            if not sheet.auto_filter.ref:
                errors.append(f"{sheet_name} auto filter 누락")
    if "채용공고" in workbook.sheetnames:
        sheet = workbook["채용공고"]
        if jobs and sheet.max_row < 2:
            errors.append("포함 공고가 있는데 채용공고 시트가 비어 있음")
        header_values = [cell.value for cell in sheet[1]]
        if header_values != JOB_HEADERS:
            errors.append("채용공고 헤더 불일치")

    for job in jobs:
        if str(job.get("직무분류")) != "Youth Intern / IT":
            continue
        title = str(job.get("공고명") or "")
        decisive_text = " ".join(
            [
                title,
                str(job.get("주요업무 요약") or ""),
                str(job.get("필수기술") or ""),
                str(job.get("지원 메모") or ""),
            ]
        )
        if contains_any(title, NON_IT_YOUTH_INTERN_TITLE_TERMS) and not contains_it_intern_term(decisive_text):
            errors.append(f"비IT 청년인턴 의심 공고 포함: {title}")
    return errors


def build_summary_message(collect_date: str, checked_count: int, jobs: Sequence[Dict[str, object]]) -> str:
    high_fit = sum(1 for job in jobs if job.get("2년차 적합도") == "상")
    urgent = sum(
        1
        for job in jobs
        if (days := days_until_deadline(job.get("마감일"), collect_date)) is not None and 0 <= days <= 7
    )
    lines = [
        f"[IT 채용공고] {collect_date} 2년차 백엔드 맞춤 수집",
        "",
        f"총 확인: {checked_count}개",
        f"엑셀 포함: {len(jobs)}개",
        f"적합도 상: {high_fit}개",
        f"마감 7일 이내: {urgent}개",
    ]
    lines += ["", "엑셀 파일을 첨부했습니다."]
    return "\n".join(lines)


def send_telegram(token: str, chat_id: str, message: str, document_path: Path) -> Tuple[bool, str]:
    if not re.fullmatch(r"it-backend-jobs-\d{4}-\d{2}-\d{2}\.xlsx", document_path.name):
        return False, f"non-final attachment name refused: {document_path.name}"
    session = requests.Session()
    message_response = session.post(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data={"chat_id": chat_id, "text": message},
        timeout=30,
    )
    if not message_response.ok:
        return False, f"sendMessage HTTP {message_response.status_code}: {message_response.text[:200]}"

    with document_path.open("rb") as document:
        document_response = session.post(
            f"https://api.telegram.org/bot{token}/sendDocument",
            data={"chat_id": chat_id},
            files={"document": (document_path.name, document, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            timeout=60,
        )
    if not document_response.ok:
        return False, f"sendDocument HTTP {document_response.status_code}: {document_response.text[:200]}"
    return True, "sendMessage/sendDocument succeeded"


def delivery_key(collect_date: str, chat_id: str) -> str:
    return f"{collect_date}:{chat_id}"


def read_delivery_state(path: Path) -> Dict[str, object]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return data if isinstance(data, dict) else {}


def delivery_already_sent(path: Path, key: str) -> bool:
    return key in read_delivery_state(path)


def record_delivery(path: Path, key: str, output_path: str, checked_at: str) -> None:
    state = read_delivery_state(path)
    state[key] = {"output_path": output_path, "sent_at": checked_at}
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def append_memory(path: Optional[Path], lines: Sequence[str]) -> None:
    if path is None:
        return
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("a", encoding="utf-8") as handle:
            handle.write("\n" + "\n".join(lines).rstrip() + "\n")
    except OSError as exc:
        print(f"memory write skipped: {exc}")


def source_detail_limit(source: str, max_details: int) -> int:
    return max_details


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--max-details", type=int, default=200)
    parser.add_argument("--sources", type=parse_source_names, default=None, help="Override the date-based source pair")
    parser.add_argument("--send-telegram", action="store_true")
    parser.add_argument("--chat-id", default=CHAT_ID)
    parser.add_argument("--token-env", default="TELEGRAM_BOT_TOKEN")
    parser.add_argument("--memory-path", type=Path, default=None)
    parser.add_argument("--research-ledger", type=Path, default=DEFAULT_RESEARCH_LEDGER)
    parser.add_argument("--delivery-state", type=Path, default=DEFAULT_DELIVERY_STATE)
    parser.add_argument("--force-send", action="store_true", help="Override the one-send-per-date Telegram guard")
    parser.add_argument("--quiet", action="store_true", help="Print compact logs without repeating the full Telegram message")
    return parser.parse_args(argv)


def main() -> int:
    args = parse_args()
    now = kst_now()
    collect_date = now.strftime("%Y-%m-%d")
    checked_at = now.strftime("%Y-%m-%d %H:%M KST")
    output_path = args.output or Path(f"/tmp/it-backend-jobs-{collect_date}.xlsx")

    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 Codex job automation"})

    source_summary = scan_official_sources(session)
    collectors = {
        "wanted": ("Wanted", collect_wanted_jobs),
        "rallit": ("Rallit", collect_rallit_jobs),
        "saramin": ("Saramin", collect_saramin_jobs),
        "jobkorea": ("JobKorea", collect_jobkorea_jobs),
        "jumpit": ("Jumpit", collect_jumpit_jobs),
        "catch": ("Catch", collect_catch_jobs),
        "incruit": ("Incruit", collect_incruit_jobs),
        "worknet": ("Worknet", collect_worknet_jobs),
        "starting": ("Starting", collect_starting_jobs),
    }
    active_sources = args.sources if args.sources is not None else scheduled_sources(collect_date)
    research_ledger = load_research_ledger(args.research_ledger)
    collected_jobs: List[Dict[str, object]] = []
    excluded: List[Dict[str, object]] = []
    detail_checked = 0
    skipped_researched = 0

    def run_collector(source: str, fallback: bool = False) -> None:
        nonlocal detail_checked, skipped_researched
        label, collector = collectors[source]
        try:
            result = collector(
                session,
                collect_date,
                checked_at,
                source_detail_limit(source, args.max_details),
                set(research_ledger),
            )
        except requests.RequestException as exc:
            excluded.append(
                {
                    "회사명": "미확인",
                    "공고명": f"{label} 목록 접근 실패",
                    "링크": "",
                    "출처": label,
                    "제외 이유": f"목록 접근 실패: {exc.__class__.__name__}",
                    "확인일시": checked_at,
                }
            )
            source_summary.append(f"{label} 목록 접근 실패: {exc.__class__.__name__}")
            return
        collected_jobs.extend(result.included)
        excluded.extend(result.excluded)
        detail_checked += result.checked
        skipped_researched += result.skipped_researched
        research_ledger.update(result.research_records)
        prefix = "보조 " if fallback else ""
        source_summary.append(f"{label} {prefix}상세 {result.checked}개 확인 / 조사 이력 제외 {result.skipped_researched}개")

    for source in active_sources:
        run_collector(source)
    source_summary.append(f"기본 수집 후 고유 포함 {selected_job_count(collected_jobs, collect_date)}개 / 목표 {TARGET_INCLUDED_JOBS}개")
    if args.sources is None and selected_job_count(collected_jobs, collect_date) < TARGET_INCLUDED_JOBS:
        source_summary.append(
            f"기본 수집 고유 포함 {selected_job_count(collected_jobs, collect_date)}개로 부족: 보조 플랫폼을 목표 {TARGET_INCLUDED_JOBS}개까지 순차 확인"
        )
        for source in FALLBACK_SOURCES:
            if selected_job_count(collected_jobs, collect_date) >= TARGET_INCLUDED_JOBS:
                break
            run_collector(source, fallback=True)
            source_summary.append(
                f"{collectors[source][0]} 보조 수집 후 고유 포함 {selected_job_count(collected_jobs, collect_date)}개 / 목표 {TARGET_INCLUDED_JOBS}개"
            )
    save_research_ledger(args.research_ledger, research_ledger)

    research_note = f"상세 조사 이력: 이번 실행 {detail_checked}개 확인 / 기존 조사 공고 {skipped_researched}개 제외"
    source_summary.append(research_note)
    jobs = select_jobs_for_workbook(collected_jobs, collect_date, limit=TARGET_INCLUDED_JOBS)
    checked_count = detail_checked

    limitation_note = (
        "공식 페이지는 접근 가능 여부와 키워드를 확인했고, 최종 포함은 각 플랫폼의 공개 API 또는 "
        "상세 HTML 본문을 확인한 공고로 제한. 정상 상세 확인을 마친 공고는 "
        "영구 조사 이력에 기록해 이후 재조사하지 않음"
    )
    build_workbook(output_path, collect_date, jobs, excluded, source_summary, limitation_note, checked_count)
    validation_errors = validate_workbook_for_delivery(output_path, jobs)
    message = build_summary_message(collect_date, checked_count, jobs)

    telegram_status = "not requested"
    if args.send_telegram:
        if validation_errors:
            telegram_status = "not sent: validation failed - " + "; ".join(validation_errors[:3])
            print(telegram_status)
        elif delivery_already_sent(args.delivery_state, delivery_key(collect_date, args.chat_id)) and not args.force_send:
            telegram_status = f"not sent: already delivered for {collect_date} to chat {args.chat_id}"
            print(telegram_status)
        else:
            token = os.environ.get(args.token_env)
            if not token:
                telegram_status = f"failed: {args.token_env} is not set"
                print(telegram_status)
            else:
                ok, telegram_status = send_telegram(token, args.chat_id, message, output_path)
                print(telegram_status)
                if ok:
                    record_delivery(args.delivery_state, delivery_key(collect_date, args.chat_id), str(output_path), checked_at)
                if not ok and telegram_status.startswith("sendDocument"):
                    fallback = f"Telegram document delivery failed. Local file: {output_path}"
                    requests.post(
                        f"https://api.telegram.org/bot{token}/sendMessage",
                        data={"chat_id": args.chat_id, "text": fallback},
                        timeout=30,
                    )
    elif validation_errors:
        print("validation failed: " + "; ".join(validation_errors[:3]))

    if args.send_telegram and validation_errors:
        memory_lines = [
            f"## {collect_date} scripted rerun blocked",
            "",
            f"- Run time: finished around {checked_at}.",
            f"- Telegram delivery: {telegram_status}.",
            f"- XLSX file: `{output_path}`.",
            f"- Included jobs before validation block: {len(jobs)}.",
            "- Validation errors:",
        ]
        for error in validation_errors[:5]:
            memory_lines.append(f"  - {error}")
        append_memory(args.memory_path, memory_lines)
        print(f"output={output_path}")
        print(f"sources={','.join(active_sources)} checked={checked_count} included={len(jobs)} excluded={len(excluded)}")
        return 2

    memory_lines = [
        f"## {collect_date} scripted rerun",
        "",
        f"- Run time: finished around {checked_at}.",
        f"- Telegram delivery: {telegram_status}.",
        f"- XLSX file: `{output_path}`.",
        f"- Included jobs: {len(jobs)}.",
        f"- {research_note}.",
    ]
    memory_lines.append("- Next run: keep using the single Python script to avoid fragmented network approvals.")
    append_memory(args.memory_path, memory_lines)

    print(f"output={output_path}")
    print(f"sources={','.join(active_sources)} checked={checked_count} included={len(jobs)} excluded={len(excluded)}")
    if not args.quiet:
        print(message)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
